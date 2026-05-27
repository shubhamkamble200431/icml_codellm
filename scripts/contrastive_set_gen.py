"""
Qwen2.5-Coder  —  Contrastive Layer-wise Activation Extraction  (MBPP+)
========================================================================
Step 1 of the CODE_LLM pipeline.

Runs N_RUNS=5 generations per problem on MBPP+, records right/wrong verdicts
and saves per-layer hidden-state activations to HDF5 files.

Contrastive filter:  0 < n_passed < N_RUNS  (n_passed ∈ {1,2,3,4} — mixed verdict)
Stratified 70/30 split applied to the contrastive problems only.
Output: ../mbppplus/<model_key>_<ts>/

Models supported:
  • Qwen2.5-Coder-1.5B-Instruct
  • Qwen2.5-Coder-7B-Instruct

Usage:
    python contrastive_set_gen.py                        # run all Qwen models on MBPP+
    python contrastive_set_gen.py --models qwen-coder-1.5b-instruct
    python contrastive_set_gen.py --skip-probing
    python contrastive_set_gen.py --recompute-split
"""

import argparse
import gc
import gzip
import json
import math
import os
import pickle
import re
import shutil
import subprocess
import sys
import tempfile
import urllib.request
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

try:
    from datasets import load_dataset as hf_load_dataset
    HF_DATASETS_OK = True
except ImportError:
    HF_DATASETS_OK = False
    print("WARNING: huggingface datasets not found — only HumanEval available."
          "  pip install datasets")

os.environ.setdefault("HF_HUB_DISABLE_XET", "1")

import h5py
import numpy as np
import torch
from tqdm import tqdm
from transformers import AutoModelForCausalLM, AutoTokenizer

try:
    from sklearn.linear_model import LogisticRegression
    from sklearn.model_selection import StratifiedKFold, cross_validate
    from sklearn.preprocessing import StandardScaler
    SKLEARN_OK = True
except ImportError:
    SKLEARN_OK = False
    print("WARNING: scikit-learn not found – probing skipped.  pip install scikit-learn")

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False
    print("WARNING: matplotlib not found – plots skipped.  pip install matplotlib")

N_RUNS         = 5
TEMPERATURE    = 0.8
TOP_P          = 0.95
MAX_NEW_TOKENS = 512
EXEC_TIMEOUT   = 10
PROBE_CV_FOLDS = 5
_MBPPPLUS_ROOT = str(Path(__file__).parent.parent / "mbppplus")
OUT_ROOT       = _MBPPPLUS_ROOT
TRAIN_RATIO    = 0.7
SPLIT_SEED     = 42

DATASET_MENU = {
    "1": ("mbppplus", "MBPP+  (~378 problems, more robust tests — evalplus)"),
}

HUMANEVAL_URL  = (
    "https://raw.githubusercontent.com/"
    "openai/human-eval/master/data/HumanEval.jsonl.gz"
)
HUMANEVAL_PATH = "HumanEval.jsonl"

BATCH_CONFIG = {
    "qwen-coder-1.5b-instruct": 4,
    "qwen-coder-7b-instruct":   1,
}

MODEL_REGISTRY = OrderedDict([
    ("qwen-coder-1.5b-instruct", {
        "hf_id":      "Qwen/Qwen2.5-Coder-1.5B-Instruct",
        "chat":       True,
        "display":    "Qwen2.5-Coder-1.5B",
        "system_msg": "You are an expert Python programmer.",
    }),
    ("qwen-coder-7b-instruct", {
        "hf_id":      "Qwen/Qwen2.5-Coder-7B-Instruct",
        "chat":       True,
        "display":    "Qwen2.5-Coder-7B",
        "system_msg": "You are an expert Python programmer.",
    }),
])

PASSK_VALUES = [1, 2, 3, 5]



def strip_code_fences(text: str) -> str:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    if normalized.startswith("```"):
        parts = normalized.split("\n", 1)
        normalized = parts[1] if len(parts) == 2 else ""
    return normalized


def last_function_signature(prompt: str) -> str | None:
    for line in reversed(prompt.splitlines()):
        stripped = line.strip()
        if stripped.startswith(("def ", "async def ")):
            return stripped
    return None


def prompt_body_indent(prompt: str) -> str:
    for line in reversed(prompt.splitlines()):
        if line.strip():
            indent_width = len(line) - len(line.lstrip(" "))
            return " " * indent_width
    return "    "


def normalize_body_indentation(lines: list[str], body_indent: str) -> str:
    nonempty = [line for line in lines if line.strip()]
    if not nonempty:
        return ""
    common_indent = min(len(line) - len(line.lstrip(" ")) for line in nonempty)
    normalized_lines: list[str] = []
    for line in lines:
        if not line.strip():
            normalized_lines.append("")
            continue
        stripped = line[common_indent:] if len(line) >= common_indent else line.lstrip(" ")
        normalized_lines.append(body_indent + stripped)
    return "\n".join(normalized_lines).rstrip()


def strip_repeated_signature_and_docstring(raw_text: str, prompt: str) -> str | None:
    signature = last_function_signature(prompt)
    if not signature:
        return None
    body_indent = prompt_body_indent(prompt)
    lines = raw_text.split("\n")
    signature_index = None
    for index, line in enumerate(lines):
        if line.strip() == signature:
            signature_index = index
            break
    if signature_index is None:
        return None
    remaining = lines[signature_index + 1:]
    while remaining and not remaining[0].strip():
        remaining = remaining[1:]
    if remaining and remaining[0].strip().startswith(('"""', "'''")):
        quote = '"""' if remaining[0].strip().startswith('"""') else "'''"
        remaining = remaining[1:]
        while remaining:
            line = remaining[0]
            remaining = remaining[1:]
            if quote in line:
                break
        while remaining and not remaining[0].strip():
            remaining = remaining[1:]
    truncated: list[str] = []
    for line in remaining:
        stripped = line.strip()
        if stripped in {"```", "<end_of_turn>", "<start_of_turn>model"}:
            break
        if line and not line.startswith((" ", "\t")):
            break
        truncated.append(line)
    return normalize_body_indentation(truncated, body_indent)


def truncate_to_function_continuation(text: str, prompt: str) -> str:
    normalized = strip_code_fences(text)
    if prompt in normalized:
        normalized = normalized.split(prompt, 1)[1]
    signature_stripped = strip_repeated_signature_and_docstring(normalized, prompt)
    if signature_stripped:
        return signature_stripped
    lines: list[str] = []
    for line in normalized.split("\n"):
        stripped = line.strip()
        if stripped in {"```", "<end_of_turn>", "<start_of_turn>model"}:
            break
        if line and not line.startswith((" ", "\t")):
            if stripped.startswith(("def ", "class ", "if __name__", "print(", "#")):
                break
        lines.append(line)
    return "\n".join(lines).rstrip()



def get_stop_token_ids(tokenizer) -> list[int]:
    stop_strings = [
        "<end_of_turn>", "<start_of_turn>",
        "<|im_end|>",    "<|im_start|>",
    ]
    ids: list[int] = []
    if tokenizer.eos_token_id is not None:
        ids.append(tokenizer.eos_token_id)
    for token in stop_strings:
        token_id = tokenizer.convert_tokens_to_ids(token)
        if token_id != tokenizer.unk_token_id and token_id not in ids:
            ids.append(token_id)
    return ids



def _user_msg(problem: dict) -> str:
    if problem.get("execution_mode") == "stdin_stdout":
        return (
            "Solve the following programming problem.\n"
            "Write a complete Python program that reads input from stdin and writes the answer to stdout.\n"
            "Return only the Python code. Do not add explanations or markdown fences.\n\n"
            f"{problem['prompt']}"
        )
    return (
        "Complete the following Python function.\n"
        "Return only Python code for the function continuation.\n"
        "Do not add explanations, markdown, comments outside the code, or tests.\n\n"
        f"{problem['prompt']}"
    )


def build_prompt(problem: dict, model_cfg: dict, tokenizer) -> str:
    if model_cfg["chat"]:
        msgs = []
        if model_cfg.get("system_msg"):
            msgs.append({"role": "system", "content": model_cfg["system_msg"]})
        msgs.append({"role": "user", "content": _user_msg(problem)})
        try:
            return tokenizer.apply_chat_template(
                msgs, tokenize=False, add_generation_prompt=True
            )
        except Exception:
            return tokenizer.apply_chat_template(
                [{"role": "user", "content": _user_msg(problem)}],
                tokenize=False, add_generation_prompt=True,
            )
    return _user_msg(problem)



def download_humaneval() -> None:
    if os.path.exists(HUMANEVAL_PATH):
        return
    print("Downloading HumanEval …")
    gz = HUMANEVAL_PATH + ".gz"
    urllib.request.urlretrieve(HUMANEVAL_URL, gz)
    with gzip.open(gz, "rb") as fi, open(HUMANEVAL_PATH, "wb") as fo:
        fo.write(fi.read())
    os.unlink(gz)
    print(f"Saved: {HUMANEVAL_PATH}")


def load_humaneval() -> list[dict]:
    rows = []
    with open(HUMANEVAL_PATH, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    print(f"Loaded {len(rows)} HumanEval problems.")
    return rows



def _require_hf(name: str) -> None:
    if not HF_DATASETS_OK:
        raise ImportError(
            f"huggingface `datasets` package required to load {name}.\n"
            "  pip install datasets"
        )


def load_humanevalplus() -> list[dict]:
    _require_hf("HumanEval+")
    print("Loading evalplus/humanevalplus …")
    ds = hf_load_dataset("evalplus/humanevalplus", split="test", trust_remote_code=True)
    rows = []
    for item in ds:
        rows.append({
            "task_id":            item["task_id"],
            "prompt":             item["prompt"],
            "entry_point":        item["entry_point"],
            "test":               item["test"],
            "execution_mode":     "function",
            "canonical_solution": item.get("canonical_solution", ""),
        })
    print(f"Loaded {len(rows)} HumanEval+ problems.")
    return rows


def _mbpp_to_normalized(item: dict, id_prefix: str = "Mbpp") -> dict:
    """Convert one MBPP-style row to the shared normalized schema."""
    code = item.get("code", "") or item.get("canonical_solution", "")
    sig_match = re.search(r"(def \w+\([^)]*\)):", code)
    if sig_match:
        sig_str   = sig_match.group(1)
        func_name = sig_str.split("(")[0][4:].strip()
    else:
        func_name = "solution"
        sig_str   = "def solution(*args)"

    prompt = f"{sig_str}:\n    \"\"\"\n    {item.get('text', item.get('prompt', ''))}\n    \"\"\"\n"

    test_list = item.get("test_list") or []
    test_code = item.get("test")

    return {
        "task_id":            f"{id_prefix}/{item['task_id']}",
        "prompt":             prompt,
        "entry_point":        func_name,
        "test":               test_code,
        "test_list":          test_list,
        "execution_mode":     "function",
        "canonical_solution": code,
    }


def load_mbpp() -> list[dict]:
    _require_hf("MBPP")
    print("Loading google-research-datasets/mbpp …")
    ds   = hf_load_dataset("google-research-datasets/mbpp", split="test")
    rows = [_mbpp_to_normalized(item) for item in ds]
    print(f"Loaded {len(rows)} MBPP problems.")
    return rows


def load_mbppplus() -> list[dict]:
    _require_hf("MBPP+")
    print("Loading evalplus/mbppplus …")
    ds   = hf_load_dataset("evalplus/mbppplus", split="test")
    rows = []
    for item in ds:
        code = item.get("canonical_solution", "") or item.get("code", "")
        sig_match = re.search(r"(def \w+\([^)]*\)):", code)
        if sig_match:
            sig_str   = sig_match.group(1)
            func_name = sig_str.split("(")[0][4:].strip()
        else:
            func_name = item.get("entry_point", "solution")
            sig_str   = f"def {func_name}(*args)"

        text   = item.get("text", item.get("prompt", ""))
        prompt = f"{sig_str}:\n    \"\"\"\n    {text}\n    \"\"\"\n"

        rows.append({
            "task_id":            f"Mbpp/{item['task_id']}",
            "prompt":             prompt,
            "entry_point":        item.get("entry_point", func_name),
            "test":               item.get("test"),
            "test_list":          item.get("test_list") or [],
            "execution_mode":     "function",
            "canonical_solution": code,
        })
    print(f"Loaded {len(rows)} MBPP+ problems.")
    return rows


def load_apps(difficulty: str = "introductory", max_problems: int = 500,
              split: str = "test") -> list[dict]:
    _require_hf("APPS")
    print(f"Loading codeparrot/apps  [difficulty={difficulty}, max={max_problems}] …")

    try:
        from huggingface_hub import hf_hub_download as _hf_dl
    except ImportError as exc:
        raise RuntimeError("pip install huggingface_hub to load APPS") from exc

    try:
        jsonl_path = _hf_dl(
            repo_id="loubnabnl/apps",
            filename=f"{split}.jsonl",
            repo_type="dataset",
        )
    except Exception as exc:
        raise RuntimeError(f"Cannot download APPS data from loubnabnl/apps: {exc}") from exc

    rows: list[dict] = []
    with open(jsonl_path, "rb") as fh:
        for line in fh:
            if max_problems and len(rows) >= max_problems:
                break
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line.decode("utf-8"))
            except Exception:
                continue

            diff = item.get("difficulty", "")
            if difficulty != "all" and diff != difficulty:
                continue

            raw_io = item.get("input_output") or "{}"
            try:
                io_dict = json.loads(raw_io) if isinstance(raw_io, str) else raw_io
            except Exception:
                io_dict = {}
            inputs  = io_dict.get("inputs",  [])
            outputs = io_dict.get("outputs", [])
            test_cases = [{"input": str(inp), "output": str(out)}
                          for inp, out in zip(inputs, outputs) if out is not None]

            prompt_text = item.get("question", "").strip()
            sc = (item.get("starter_code") or "").strip()
            if sc:
                prompt_text += f"\n\nStarter code:\n{sc}"

            rows.append({
                "task_id":        f"apps/{item['id']}",
                "prompt":         prompt_text,
                "entry_point":    None,
                "test":           None,
                "test_list":      None,
                "test_cases":     test_cases,
                "execution_mode": "stdin_stdout",
                "difficulty":     diff,
            })

    print(f"Loaded {len(rows)} APPS problems.")
    return rows


def load_livecodebench(version_tag: str = "v5") -> list[dict]:
    _require_hf("LiveCodeBench")
    print(f"Loading livecodebench/code_generation_lite [{version_tag}] …")

    try:
        from huggingface_hub import snapshot_download as _snap
    except ImportError as exc:
        raise RuntimeError("pip install huggingface_hub to load LiveCodeBench") from exc

    try:
        cache_dir = _snap(
            "livecodebench/code_generation_lite",
            repo_type="dataset",
            ignore_patterns=["*.py"],
        )
    except Exception as exc:
        raise RuntimeError(
            f"Cannot download livecodebench/code_generation_lite: {exc}"
        ) from exc

    import glob as _glob
    parquet_files = sorted(_glob.glob(
        os.path.join(cache_dir, "**", "*.parquet"), recursive=True))
    jsonl_files   = sorted(_glob.glob(
        os.path.join(cache_dir, "**", "*.jsonl"),   recursive=True))

    if parquet_files:
        ds = hf_load_dataset("parquet", data_files={"test": parquet_files}, split="test")
    elif jsonl_files:
        ds = hf_load_dataset("json",    data_files={"test": jsonl_files},    split="test")
    else:
        raise RuntimeError(
            "No parquet/jsonl files found in livecodebench/code_generation_lite cache"
        )

    if "release_version" in ds.column_names:
        ds = ds.filter(lambda x: x.get("release_version") == version_tag)
    elif "version" in ds.column_names:
        ds = ds.filter(lambda x: x.get("version") == version_tag)

    rows: list[dict] = []
    for item in ds:
        raw_tc = item.get("public_test_cases") or []
        test_cases: list[dict] = []
        for tc in raw_tc:
            if isinstance(tc, str):
                try:
                    tc = json.loads(tc)
                except Exception:
                    continue
            if not isinstance(tc, dict):
                continue
            test_cases.append({
                "input":  str(tc.get("input",  "")),
                "output": str(tc.get("output", "")),
            })

        raw_meta = item.get("metadata") or {}
        if isinstance(raw_meta, str):
            try:
                raw_meta = json.loads(raw_meta)
            except Exception:
                raw_meta = {}
        meta    = raw_meta if isinstance(raw_meta, dict) else {}
        starter = (meta.get("starter_code") or item.get("starter_code") or "").strip()
        content     = item.get("question_content", item.get("question", "")).strip()
        title       = item.get("question_title", "")
        prompt_text = f"Problem: {title}\n\n{content}"
        if starter:
            prompt_text += f"\n\nStarter code:\n{starter}"

        rows.append({
            "task_id":        f"lcb/{item.get('question_id', item.get('id', len(rows)))}",
            "prompt":         prompt_text,
            "entry_point":    None,
            "test":           None,
            "test_list":      None,
            "test_cases":     test_cases,
            "execution_mode": "stdin_stdout",
        })
    print(f"Loaded {len(rows)} LiveCodeBench problems.")
    return rows



def extract_full_program(text: str) -> str:
    """Extract a complete Python program from model output (stdin/stdout tasks)."""
    text = text.strip()
    m = re.search(r"```python\s*\n(.*?)```", text, re.DOTALL)
    if m:
        return m.group(1).strip()
    m = re.search(r"```\s*\n(.*?)```", text, re.DOTALL)
    if m:
        return m.group(1).strip()
    lines = text.split("\n")
    for i, line in enumerate(lines):
        s = line.strip()
        if s.startswith(("import ", "from ", "def ", "class ", "n ", "n=",
                          "input", "print", "sys.", "for ", "while ", "if ")):
            return "\n".join(lines[i:]).rstrip()
    return text


def make_split(problems_indexed):
    """Deterministic 70/30 split on whatever list of (idx, problem) pairs is given."""
    import random
    rng = random.Random(SPLIT_SEED)
    shuffled = list(problems_indexed)
    rng.shuffle(shuffled)
    cut        = int(len(shuffled) * TRAIN_RATIO)
    train_list = shuffled[:cut]
    test_list  = shuffled[cut:]
    train_ids  = {p["task_id"] for _, p in train_list}
    test_ids   = {p["task_id"] for _, p in test_list}
    return train_list, test_list, train_ids, test_ids



def filter_contrastive(
    summary: list[dict],
    problems_indexed: list[tuple],
) -> tuple[list[tuple], list[dict], list[dict], list[dict]]:
    """
    Keep only problems with 0 < n_passed < N_RUNS  (i.e. n_passed ∈ {1, 2, 3, 4} out of 5).
    This ensures genuine ambiguity: the model neither always passes nor always fails,
    and at least 2 runs went each way.

    Returns
    -------
    contrastive_indexed : list[(idx, problem)]  — filtered subset of problems_indexed
    contrastive_summary : list[dict]            — matching summary rows
    all_pass_summary    : list[dict]            — problems where n_passed == N_RUNS (all right)
    all_fail_summary    : list[dict]            — problems where n_passed == 0 (all wrong)
    """
    summary_by_id = {s["task_id"]: s for s in summary}

    contrastive_indexed: list[tuple] = []
    contrastive_summary: list[dict]  = []
    all_pass_summary:    list[dict]  = []
    all_fail_summary:    list[dict]  = []

    for idx, prob in problems_indexed:
        s = summary_by_id.get(prob["task_id"])
        if s is None:
            continue
        n = s["n_passed"]
        if 0 < n < N_RUNS:
            contrastive_indexed.append((idx, prob))
            contrastive_summary.append(s)
        elif n == N_RUNS:
            all_pass_summary.append(s)
        else:
            all_fail_summary.append(s)

    return contrastive_indexed, contrastive_summary, all_pass_summary, all_fail_summary



def _contrastive_section(
    lines: list,
    label: str,
    split_summary: list[dict],
    contrastive_summary: list[dict],
    all_pass_summary: list[dict],
    all_fail_summary: list[dict],
    SEP2: str,
) -> None:
    """Append one split's contrastive overview + per-prompt table to lines."""
    n_split       = len(split_summary)
    n_contrastive = len(contrastive_summary)
    n_all_pass    = len(all_pass_summary)
    n_all_fail    = len(all_fail_summary)
    pct           = n_contrastive / n_split * 100 if n_split else 0.0

    tr  = sum(s["n_passed"] for s in contrastive_summary)
    tw  = sum(s["n_failed"] for s in contrastive_summary)
    tot = tr + tw

    lines.append(f"  {label.upper()} OVERVIEW")
    lines.append(SEP2)
    lines.append(f"  Problems (total)                  : {n_split}")
    lines.append(f"  All-pass  (excluded)              : {n_all_pass:>4}  "
                 f"({n_all_pass/n_split*100:5.1f}%)")
    lines.append(f"  All-fail  (excluded)              : {n_all_fail:>4}  "
                 f"({n_all_fail/n_split*100:5.1f}%)")
    lines.append(f"  Contrastive / mixed (kept)        : {n_contrastive:>4}  "
                 f"({pct:5.1f}%)   ← {n_contrastive}/{n_split}")
    lines.append("")

    lines.append(f"  AGGREGATE — CONTRASTIVE {label.upper()}")
    lines.append(SEP2)
    lines.append(f"  Total runs        : {tot}  "
                 f"({n_contrastive} problems × {N_RUNS} runs)")
    if tot > 0:
        lines.append(f"  Right runs (pass) : {tr:>5}  ({tr/tot*100:5.1f}%)")
        lines.append(f"  Wrong runs (fail) : {tw:>5}  ({tw/tot*100:5.1f}%)")
        lines.append(f"  Right/Wrong ratio : {tr/max(tw,1):.3f}")
    lines.append("")

    lines.append(f"  RUN DISTRIBUTION — CONTRASTIVE {label.upper()}")
    lines.append(SEP2)
    lines.append(f"  {'n_passed':>9}  {'n_problems':>10}  {'%':>7}  bar")
    lines.append(SEP2)
    for k in range(1, N_RUNS):
        cnt   = sum(1 for s in contrastive_summary if s["n_passed"] == k)
        pct_k = cnt / n_contrastive * 100 if n_contrastive else 0.0
        bar   = "█" * int(pct_k / 2)
        lines.append(f"  {k:>4}/{N_RUNS} runs  {cnt:>10}  {pct_k:>6.1f}%  {bar}")
    lines.append("")

    lines.append(f"  PER-PROMPT BREAKDOWN — CONTRASTIVE {label.upper()}")
    lines.append(SEP2)
    col = max((len(s["task_id"]) for s in contrastive_summary), default=16)
    lines.append(
        f"  {'task_id':<{col}}  {'right':>6}  {'wrong':>6}  {'total':>6}  "
        f"{'right%':>7}  run pattern"
    )
    lines.append(SEP2)
    for s in sorted(contrastive_summary, key=lambda x: x["task_id"]):
        r_runs  = s["n_passed"]
        w_runs  = s["n_failed"]
        pattern = "■" * r_runs + "□" * w_runs
        lines.append(
            f"  {s['task_id']:<{col}}  {r_runs:>6}  {w_runs:>6}  {N_RUNS:>6}  "
            f"{r_runs/N_RUNS*100:>6.1f}%  {pattern}"
        )
    lines.append("")


def report_contrastive_stats(
    train_summary: list[dict],
    contrastive_train_summary: list[dict],
    train_all_pass_summary: list[dict],
    train_all_fail_summary: list[dict],
    test_summary: list[dict],
    contrastive_test_summary: list[dict],
    test_all_pass_summary: list[dict],
    test_all_fail_summary: list[dict],
    out_dir: str,
    display: str,
) -> None:
    """
    Prints and saves contrastive_stats.txt with train and test sections:
      • Overview counts (total / all-pass / all-fail / contrastive) per split
      • Aggregate right/wrong run counts per split
      • Run distribution histogram per split
      • Per-prompt table per split
    """
    SEP  = "=" * 72
    SEP2 = "-" * 72
    lines: list[str] = []

    lines.append(SEP)
    lines.append(f"  CONTRASTIVE STATS  —  {display}")
    lines.append(f"  Contrastive filter applied to BOTH train and test splits")
    lines.append(SEP)
    lines.append("")

    _contrastive_section(
        lines, "TRAIN SET",
        train_summary, contrastive_train_summary,
        train_all_pass_summary, train_all_fail_summary,
        SEP2,
    )

    lines.append(SEP2)
    lines.append("")

    _contrastive_section(
        lines, "TEST SET",
        test_summary, contrastive_test_summary,
        test_all_pass_summary, test_all_fail_summary,
        SEP2,
    )

    lines.append(SEP)

    text = "\n".join(lines)
    tqdm.write(text)

    stats_path = Path(out_dir) / "contrastive_stats.txt"
    stats_path.write_text(text, encoding="utf-8")
    tqdm.write(f"  Stats  → {stats_path}")



def save_h5(path: Path, arr: np.ndarray) -> None:
    with h5py.File(path, "w") as hf:
        hf.create_dataset("activation", data=arr.astype(np.float32),
                          compression="gzip", compression_opts=4)


def load_h5(path: Path) -> np.ndarray:
    with h5py.File(path, "r") as hf:
        return hf["activation"][:]



def pass_at_k(n: int, c: int, k: int) -> float:
    if k > n:
        return float("nan")
    if n - c < k:
        return 1.0
    result = 1.0
    for i in range(k):
        result *= (n - c - i) / (n - i)
    return 1.0 - result


def _agg_passk(summary_rows: list[dict]) -> dict:
    out = {}
    for k in PASSK_VALUES:
        key  = f"pass@{k}"
        vals = [s[key] for s in summary_rows if key in s and not (s[key] != s[key])]
        out[k] = float(np.mean(vals)) if vals else 0.0
    return out



def _run_subprocess(code: str, timeout: int = EXEC_TIMEOUT):
    with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
        f.write(code)
        fname = f.name
    try:
        proc = subprocess.run(
            [sys.executable, fname],
            capture_output=True, text=True, timeout=timeout,
        )
        return proc.returncode == 0, proc.stderr
    except subprocess.TimeoutExpired:
        return False, "TIMEOUT"
    except Exception as e:
        return False, str(e)
    finally:
        try:
            os.unlink(fname)
        except Exception:
            pass


def _eval_stdin_stdout(solution_code: str, problem: dict) -> bool:
    """Run solution_code as a subprocess for each test case, compare stdout."""
    test_cases = problem.get("test_cases") or []
    if not test_cases:
        return False

    with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
        f.write(solution_code)
        fname = f.name

    try:
        for tc in test_cases:
            inp      = str(tc.get("input",  ""))
            expected = str(tc.get("output", "")).strip()
            try:
                proc = subprocess.run(
                    [sys.executable, fname],
                    input=inp,
                    capture_output=True, text=True,
                    timeout=EXEC_TIMEOUT,
                )
                if proc.returncode != 0:
                    return False
                if proc.stdout.strip() != expected:
                    return False
            except subprocess.TimeoutExpired:
                return False
            except Exception:
                return False
    finally:
        try:
            os.unlink(fname)
        except Exception:
            pass

    return True


def evaluate_all_cases(solution_code: str, problem: dict) -> bool:
    mode = problem.get("execution_mode", "function")

    if mode == "stdin_stdout":
        return _eval_stdin_stdout(solution_code, problem)

    if problem.get("test_list"):
        script = solution_code + "\n\n" + "\n".join(problem["test_list"]) + "\n"
    else:
        script = (
            f"{solution_code}\n\n"
            f"{problem['test']}\n\n"
            f"check({problem['entry_point']})\n"
        )
    ok, err = _run_subprocess(script)
    return ok and not err.strip()



def load_model(model_cfg: dict, device: str):
    hf_id    = model_cfg["hf_id"]
    quantize = model_cfg.get("quantize")
    dtype    = torch.bfloat16 if device == "cuda" else torch.float32
    print(f"\n  Loading  : {model_cfg['display']}  ({hf_id})"
          + (f"  [{quantize} quant]" if quantize else ""))

    tok = AutoTokenizer.from_pretrained(hf_id, trust_remote_code=True)
    if tok.pad_token is None:
        tok.pad_token = tok.eos_token
    tok.padding_side = "left"

    load_kwargs = dict(
        device_map="auto" if device == "cuda" else None,
        trust_remote_code=True,
    )

    if quantize and device == "cuda":
        try:
            from transformers import BitsAndBytesConfig
            quant_skip = model_cfg.get("quant_skip") or []
            if quantize == "4bit":
                bnb_cfg = BitsAndBytesConfig(
                    load_in_4bit=True,
                    bnb_4bit_quant_type="nf4",
                    bnb_4bit_compute_dtype=torch.float16,
                    bnb_4bit_use_double_quant=True,
                    llm_int8_skip_modules=quant_skip or None,
                )
            elif quantize == "8bit":
                bnb_cfg = BitsAndBytesConfig(
                    load_in_8bit=True,
                    llm_int8_compute_dtype=torch.float16,
                    llm_int8_skip_modules=quant_skip or None,
                )
            else:
                raise ValueError(f"Unknown quantize value: {quantize!r}")
            load_kwargs["quantization_config"] = bnb_cfg
        except ImportError:
            print("  WARNING: bitsandbytes not found — falling back to bfloat16.")
            load_kwargs["dtype"] = dtype
    else:
        load_kwargs["dtype"] = dtype

    mdl = AutoModelForCausalLM.from_pretrained(hf_id, **load_kwargs)
    mdl.eval()
    n = sum(p.numel() for p in mdl.parameters()) / 1e9
    if torch.cuda.is_available():
        used  = torch.cuda.memory_allocated() / 1024**3
        total = torch.cuda.get_device_properties(0).total_memory / 1024**3
        print(f"  Loaded   : {n:.2f}B params  |  GPU {used:.1f}/{total:.1f} GB")
    else:
        print(f"  Loaded   : {n:.2f}B params")
    return mdl, tok


def unload_model(mdl, tok) -> None:
    del mdl, tok
    gc.collect()
    if torch.cuda.is_available():
        try:
            torch.cuda.empty_cache()
        except Exception:
            pass
    print("  Unloaded.")



@torch.no_grad()
def generate_and_extract_batch(
    model,
    tokenizer,
    problems_batch: list[dict],
    model_cfg: dict,
) -> list[list[tuple]]:
    B         = len(problems_batch)
    first_dev = next(model.parameters()).device
    stop_ids  = get_stop_token_ids(tokenizer)

    prompts          = [build_prompt(p, model_cfg, tokenizer) for p in problems_batch]
    expanded_prompts = [p for p in prompts for _ in range(N_RUNS)]

    enc = tokenizer(
        expanded_prompts,
        return_tensors="pt",
        padding=True,
        truncation=True,
        max_length=1024,
    )
    prompt_lens    = enc["attention_mask"].sum(dim=1).tolist()
    max_prompt_len = enc["input_ids"].shape[1]

    input_ids      = enc["input_ids"].to(first_dev)
    attention_mask = enc["attention_mask"].to(first_dev)

    out_ids = model.generate(
        input_ids=input_ids,
        attention_mask=attention_mask,
        max_new_tokens=MAX_NEW_TOKENS,
        temperature=TEMPERATURE,
        top_p=TOP_P,
        do_sample=True,
        num_return_sequences=1,
        pad_token_id=tokenizer.eos_token_id,
        eos_token_id=stop_ids,
    )

    total_len    = out_ids.shape[1]
    stop_ids_set = set(stop_ids)
    out_ids_list = out_ids.cpu().tolist()

    last_positions: list[int] = []
    for i in range(B * N_RUNS):
        last_pos = total_len - 1
        for j, tok in enumerate(out_ids_list[i][max_prompt_len:]):
            if tok in stop_ids_set:
                stop_abs = max_prompt_len + j
                last_pos = max(stop_abs - 1, max_prompt_len)
                break
        last_positions.append(last_pos)

    fwd_attn = torch.zeros(B * N_RUNS, total_len, dtype=torch.long, device=first_dev)
    for i in range(B * N_RUNS):
        prompt_start = max_prompt_len - int(prompt_lens[i])
        fwd_attn[i, prompt_start : last_positions[i] + 1] = 1

    fwd = model(
        input_ids=out_ids,
        attention_mask=fwd_attn,
        output_hidden_states=True,
        return_dict=True,
    )

    positions_t    = torch.tensor(last_positions, dtype=torch.long, device=first_dev)
    seq_layer_acts = [{} for _ in range(B * N_RUNS)]
    _warned_shapes: set[tuple] = set()

    for li, hs in enumerate(fwd.hidden_states):
        if hs.dim() == 3:
            hs3 = hs
        elif hs.dim() == 4:
            shape_key = tuple(hs.shape)
            if shape_key not in _warned_shapes:
                tqdm.write(
                    f"  [hidden_states] layer {li}: 4-D {tuple(hs.shape)} "
                    f"-> taking slot 0 (Gemma-3n AltUp)"
                )
                _warned_shapes.add(shape_key)
            hs3 = hs[0]
        else:
            tqdm.write(f"  [hidden_states] layer {li}: unexpected {hs.dim()}-D, skipping")
            continue

        idx       = positions_t.view(-1, 1, 1).expand(-1, 1, hs3.shape[-1])
        last_acts = hs3.gather(1, idx).squeeze(1).float().cpu().numpy()
        for i in range(B * N_RUNS):
            seq_layer_acts[i][li] = last_acts[i]

    out_ids_cpu = out_ids.cpu()
    del fwd, out_ids
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

    results: list[list] = [[None] * N_RUNS for _ in range(B)]
    for b in range(B):
        problem = problems_batch[b]
        for r in range(N_RUNS):
            seq_i = b * N_RUNS + r
            raw = tokenizer.decode(
                out_ids_cpu[seq_i, max_prompt_len : last_positions[seq_i] + 1],
                skip_special_tokens=True,
            )
            if problem.get("execution_mode") == "stdin_stdout":
                full_sol = extract_full_program(raw)
            else:
                completion = truncate_to_function_continuation(raw, problem["prompt"])
                full_sol   = problem["prompt"] + completion
            results[b][r] = (raw, full_sol, seq_layer_acts[seq_i])

    return results



def run_analysis(
    model_key: str,
    model_cfg: dict,
    problems: list,
    device: str,
    out_dir: str,
) -> tuple[list[dict], int]:
    out_path   = Path(out_dir)
    batch_size = BATCH_CONFIG.get(model_key, 1)
    model, tokenizer = load_model(model_cfg, device)

    problem_list = list(problems)
    n_batches    = math.ceil(len(problem_list) / batch_size)

    summary: list[dict] = []
    n_layers: int | None = None

    outer = tqdm(
        range(n_batches),
        desc=f"  [{model_cfg['display']}]",
        unit="batch", position=1, leave=True,
    )

    for batch_i in outer:
        batch      = problem_list[batch_i * batch_size : (batch_i + 1) * batch_size]
        probs_only = [p for _, p in batch]

        try:
            batch_results = generate_and_extract_batch(
                model, tokenizer, probs_only, model_cfg
            )
        except torch.cuda.OutOfMemoryError:
            torch.cuda.empty_cache()
            tqdm.write(
                f"  [OOM] batch {batch_i} (size {len(batch)}), "
                "retrying one problem at a time …"
            )
            batch_results = []
            for prob in probs_only:
                try:
                    single = generate_and_extract_batch(
                        model, tokenizer, [prob], model_cfg
                    )
                    batch_results.append(single[0])
                except Exception as e:
                    tqdm.write(f"  [GEN ERR fallback] {prob['task_id']}: {e}")
                    empty: dict = {}
                    batch_results.append(
                        [("", prob["prompt"] + "    raise NotImplementedError()", empty)]
                        * N_RUNS
                    )

        for b_local, (_, problem) in enumerate(batch):
            task_safe   = problem["task_id"].replace("/", "_")
            run_records = []

            for run_idx in range(N_RUNS):
                _, full_sol, layer_acts = batch_results[b_local][run_idx]

                if n_layers is None and layer_acts:
                    n_layers = len(layer_acts)

                try:
                    passed = evaluate_all_cases(full_sol, problem)
                except Exception:
                    passed = False

                verdict = "right" if passed else "wrong"
                run_dir = out_path / task_safe / verdict / f"run{run_idx + 1}"
                run_dir.mkdir(parents=True, exist_ok=True)

                (run_dir / "generated_code.txt").write_text(full_sol, encoding="utf-8")
                for li, act in layer_acts.items():
                    save_h5(run_dir / f"layer_{li:02d}.h5", act)

                run_records.append({"run": run_idx + 1, "passed": passed})

            pass_flags = [r["passed"] for r in run_records]
            c = sum(pass_flags)
            entry: dict = {
                "task_id":    problem["task_id"],
                "n_passed":   c,
                "n_failed":   N_RUNS - c,
                "all_passed": all(pass_flags),
                "any_passed": any(pass_flags),
            }
            for k in PASSK_VALUES:
                entry[f"pass@{k}"] = pass_at_k(N_RUNS, c, k)
            summary.append(entry)

        n_right = sum(1 for s in summary if s["any_passed"])
        outer.set_postfix(right=n_right, done=len(summary))

    unload_model(model, tokenizer)
    return summary, (n_layers or 0)



def _collect_layer_buffers(out_dir: str, task_ids: set) -> dict:
    _safe = {tid.replace("/", "_", 1): tid for tid in task_ids}

    search_root = Path(out_dir)
    if (search_root / "all").is_dir():
        search_root = search_root / "all"

    bufs: dict = {}
    for prob_dir in sorted(search_root.iterdir()):
        if not prob_dir.is_dir():
            continue
        task_id = _safe.get(prob_dir.name)
        if task_id is None:
            continue
        for verdict in ("right", "wrong"):
            lbl  = 1 if verdict == "right" else 0
            vdir = prob_dir / verdict
            if not vdir.exists():
                continue
            for run_dir in sorted(vdir.iterdir()):
                if not run_dir.is_dir():
                    continue
                for h5f in sorted(run_dir.glob("layer_*.h5")):
                    li = int(h5f.stem.split("_")[1])
                    try:
                        act = load_h5(h5f)
                    except Exception as e:
                        tqdm.write(f"  [H5 ERR] {h5f}: {e}")
                        continue
                    if li not in bufs:
                        bufs[li] = {"X": [], "y": []}
                    bufs[li]["X"].append(act)
                    bufs[li]["y"].append(lbl)
    return bufs


def run_linear_probing(
    out_dir: str,
    train_ids: set,
    test_ids: set,
) -> list[dict]:
    if not SKLEARN_OK:
        tqdm.write("  scikit-learn missing — skipping probing.")
        return []

    probe_dir = Path(out_dir) / "probing"
    probe_dir.mkdir(exist_ok=True)

    tqdm.write("  Collecting train activations …")
    train_bufs = _collect_layer_buffers(out_dir, train_ids)
    tqdm.write("  Collecting test activations …")
    test_bufs  = _collect_layer_buffers(out_dir, test_ids)

    if not train_bufs:
        tqdm.write("  No train activation files found — skipping probing.")
        return []

    all_layers = sorted(train_bufs.keys())
    n_tr = len(next(iter(train_bufs.values()))["X"])
    n_te = len(next(iter(test_bufs.values()))["X"]) if test_bufs else 0
    tqdm.write(
        f"  {len(all_layers)} layers  |  "
        f"train ~{n_tr} samples  |  test ~{n_te} samples"
    )

    from sklearn.metrics import balanced_accuracy_score, accuracy_score, roc_auc_score

    results: list[dict] = []
    pbar = tqdm(
        all_layers, desc="  Probing",
        unit="layer", position=1, leave=True,
    )
    for li in pbar:
        X_tr_list = train_bufs[li]["X"]
        y_tr_list = train_bufs[li]["y"]

        if len(X_tr_list) < 6 or len(set(y_tr_list)) < 2:
            continue

        X_train = np.array(X_tr_list, dtype=np.float32)
        y_train = np.array(y_tr_list, dtype=np.int32)

        scaler  = StandardScaler()
        X_tr_sc = scaler.fit_transform(X_train)

        n_splits = min(PROBE_CV_FOLDS, int(np.bincount(y_train).min()))
        n_splits = max(n_splits, 2)

        clf = LogisticRegression(
            max_iter=1000, C=1.0, class_weight="balanced", solver="lbfgs"
        )
        cv     = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
        cv_res = cross_validate(
            clf, X_tr_sc, y_train, cv=cv,
            scoring=["balanced_accuracy", "accuracy", "roc_auc"],
            return_train_score=False,
        )

        clf.fit(X_tr_sc, y_train)
        layer_name = f"layer_{li:02d}"
        with open(probe_dir / f"{layer_name}_probe.pkl", "wb") as f:
            pickle.dump({"probe": clf, "scaler": scaler,
                         "layer_idx": li, "n_train": len(y_train)}, f)

        rec: dict = {
            "layer_idx":       li,
            "layer_name":      layer_name,
            "n_train_right":   int(np.sum(y_train == 1)),
            "n_train_wrong":   int(np.sum(y_train == 0)),
            "n_train_total":   len(y_train),
            "cv_bal_acc_mean": float(cv_res["test_balanced_accuracy"].mean()),
            "cv_bal_acc_std":  float(cv_res["test_balanced_accuracy"].std()),
            "cv_acc_mean":     float(cv_res["test_accuracy"].mean()),
            "cv_auc_mean":     float(cv_res["test_roc_auc"].mean()),
        }

        X_te_list = test_bufs.get(li, {}).get("X", [])
        y_te_list = test_bufs.get(li, {}).get("y", [])
        if len(X_te_list) >= 2 and len(set(y_te_list)) == 2:
            X_test  = np.array(X_te_list, dtype=np.float32)
            y_test  = np.array(y_te_list, dtype=np.int32)
            X_te_sc = scaler.transform(X_test)
            y_pred  = clf.predict(X_te_sc)
            y_prob  = clf.predict_proba(X_te_sc)[:, 1]
            rec.update({
                "n_test_right": int(np.sum(y_test == 1)),
                "n_test_wrong": int(np.sum(y_test == 0)),
                "n_test_total": len(y_test),
                "test_bal_acc": float(balanced_accuracy_score(y_test, y_pred)),
                "test_acc":     float(accuracy_score(y_test, y_pred)),
                "test_auc":     float(roc_auc_score(y_test, y_prob)),
            })
        else:
            rec.update({
                "n_test_right": 0, "n_test_wrong": 0, "n_test_total": 0,
                "test_bal_acc": None, "test_acc": None, "test_auc": None,
            })

        results.append(rec)
        pbar.set_postfix(
            layer=layer_name,
            tr_cv=f"{rec['cv_bal_acc_mean']:.3f}",
            te=f"{rec['test_bal_acc'] or 0:.3f}",
        )

    with open(probe_dir / "probing_results.json", "w") as f:
        json.dump(results, f, indent=2)

    if results:
        import csv
        keys = list(results[0].keys())
        with open(probe_dir / "probing_summary.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(results)

    tqdm.write(f"  Probing done : {len(results)} layers  →  {probe_dir}")
    return results



def run_5fold_cv_all(
    out_dir: str,
    all_contrastive_ids: set,
    model_display: str,
) -> list[dict]:
    """
    Run proper 5-fold stratified cross-validation over EVERY contrastive problem
    (train + test combined).  No separate holdout — each fold acts as test in turn.

    Results saved to  <out_dir>/probing_5fold_cv/
      probing_results.json   — per-layer metrics
      probing_summary.csv    — same as TSV
      cv5fold_report.txt     — human-readable text report
    """
    if not SKLEARN_OK:
        tqdm.write("  scikit-learn missing — skipping 5-fold CV.")
        return []

    n_all = len(all_contrastive_ids)
    if n_all < 5:
        tqdm.write(
            f"  WARNING: only {n_all} contrastive problems total — "
            "5-fold CV not possible, skipping."
        )
        return []

    cv_dir   = Path(out_dir) / "probing_5fold_cv"
    cv_done  = cv_dir / "probing_results.json"
    if cv_done.exists():
        tqdm.write(f"  5-fold CV already done — loading {cv_done}")
        with open(cv_done) as f:
            return json.load(f)

    cv_dir.mkdir(exist_ok=True)

    tqdm.write(f"  5-fold CV: collecting activations for {n_all} contrastive problems …")
    all_bufs = _collect_layer_buffers(out_dir, all_contrastive_ids)

    if not all_bufs:
        tqdm.write("  No activation files found for 5-fold CV — skipping.")
        return []

    all_layers = sorted(all_bufs.keys())
    n_samples  = len(next(iter(all_bufs.values()))["X"])

    tqdm.write(
        f"  {len(all_layers)} layers  |  "
        f"{n_samples} total samples  "
        f"({n_all} problems × {N_RUNS} runs)"
    )

    results: list[dict] = []
    pbar = tqdm(
        all_layers, desc="  5-fold CV",
        unit="layer", position=1, leave=True,
    )

    for li in pbar:
        X_list = all_bufs[li]["X"]
        y_list = all_bufs[li]["y"]

        if len(X_list) < 10 or len(set(y_list)) < 2:
            continue

        X = np.array(X_list, dtype=np.float32)
        y = np.array(y_list, dtype=np.int32)

        n_splits = min(PROBE_CV_FOLDS, int(np.bincount(y).min()))
        n_splits = max(n_splits, 2)

        clf    = LogisticRegression(
            max_iter=1000, C=1.0, class_weight="balanced", solver="lbfgs"
        )
        cv_obj = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
        cv_res = cross_validate(
            clf, StandardScaler().fit_transform(X), y,
            cv=cv_obj,
            scoring=["balanced_accuracy", "accuracy", "roc_auc"],
            return_train_score=False,
        )

        layer_name = f"layer_{li:02d}"
        rec: dict = {
            "layer_idx":        li,
            "layer_name":       layer_name,
            "n_samples":        len(y),
            "n_right":          int(np.sum(y == 1)),
            "n_wrong":          int(np.sum(y == 0)),
            "n_folds":          n_splits,
            "cv_bal_acc_mean":  float(cv_res["test_balanced_accuracy"].mean()),
            "cv_bal_acc_std":   float(cv_res["test_balanced_accuracy"].std()),
            "cv_acc_mean":      float(cv_res["test_accuracy"].mean()),
            "cv_acc_std":       float(cv_res["test_accuracy"].std()),
            "cv_auc_mean":      float(cv_res["test_roc_auc"].mean()),
            "cv_auc_std":       float(cv_res["test_roc_auc"].std()),
            "cv_bal_acc_folds": [round(v, 4) for v in cv_res["test_balanced_accuracy"].tolist()],
            "cv_auc_folds":     [round(v, 4) for v in cv_res["test_roc_auc"].tolist()],
        }
        results.append(rec)
        pbar.set_postfix(
            layer=layer_name,
            bal_acc=f"{rec['cv_bal_acc_mean']:.3f}±{rec['cv_bal_acc_std']:.3f}",
        )

    if not results:
        tqdm.write("  5-fold CV: no layers with sufficient samples.")
        return []

    with open(cv_dir / "probing_results.json", "w") as f:
        json.dump(results, f, indent=2)

    import csv
    keys = list(results[0].keys())
    with open(cv_dir / "probing_summary.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(results)

    SEP1 = "=" * 90
    SEP2 = "-" * 90
    CHANCE = 0.5

    best_cv  = max(results, key=lambda r: r["cv_bal_acc_mean"])
    best_auc = max(results, key=lambda r: r["cv_auc_mean"])

    lines: list[str] = []
    lines += [
        SEP1,
        f"  5-FOLD CV OVER ALL CONTRASTIVE PROBLEMS  —  {model_display}",
        f"  Generated : {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}",
        SEP1,
        "",
        f"  Total contrastive problems   : {n_all}",
        f"  Total samples (× {N_RUNS} runs)       : {n_samples}",
        f"  Layers probed                : {len(results)}",
        f"  CV folds (actual)            : {results[0]['n_folds']}",
        f"  Class balance (first layer)  : "
        f"right={results[0]['n_right']}  wrong={results[0]['n_wrong']}",
        "",
        f"  Best layer (CV bal-acc) : {best_cv['layer_name']}  "
        f"→  {best_cv['cv_bal_acc_mean']:.4f} ± {best_cv['cv_bal_acc_std']:.4f}  "
        f"(Δchance {best_cv['cv_bal_acc_mean']-CHANCE:+.4f})",
        f"  Best layer (CV AUC)    : {best_auc['layer_name']}  "
        f"→  {best_auc['cv_auc_mean']:.4f} ± {best_auc['cv_auc_std']:.4f}",
        "",
        SEP2,
        "  NOTE: all samples (train + test splits combined) used for CV.",
        "  Each fold tests on ~1/5 of total samples — no separate holdout.",
        SEP2,
        "",
    ]

    W = [10, 10, 16, 16, 14, 14, 10]
    def _row(*cells):
        return "  " + "  ".join(str(c).ljust(w) for c, w in zip(cells, W))

    lines.append(_row("Layer", "Samples", "CV bal_acc", "CV ± std",
                       "CV AUC", "CV AUC ± std", "Flag"))
    lines.append("  " + SEP2[:sum(W) + 2*len(W)])

    for rec in results:
        flags = []
        if rec["layer_name"] == best_cv["layer_name"]:
            flags.append("*best_acc")
        if rec["layer_name"] == best_auc["layer_name"]:
            flags.append("*best_auc")
        lines.append(_row(
            rec["layer_name"],
            rec["n_samples"],
            f"{rec['cv_bal_acc_mean']:.4f}",
            f"±{rec['cv_bal_acc_std']:.4f}",
            f"{rec['cv_auc_mean']:.4f}",
            f"±{rec['cv_auc_std']:.4f}",
            " ".join(flags),
        ))

    lines += ["",
              "  *best_acc = highest mean CV balanced accuracy across folds",
              "  *best_auc = highest mean CV ROC-AUC across folds",
              ""]

    lines += [
        SEP2,
        f"  PER-FOLD BREAKDOWN — best layer  ({best_cv['layer_name']})",
        SEP2,
    ]
    for fi, (ba, auc) in enumerate(
        zip(best_cv["cv_bal_acc_folds"], best_cv["cv_auc_folds"]), 1
    ):
        bar = "█" * int(ba * 20)
        lines.append(f"  Fold {fi}  bal_acc={ba:.4f}  AUC={auc:.4f}  {bar}")

    lines += ["", SEP1, "  END OF 5-FOLD CV REPORT", SEP1]

    report_text = "\n".join(lines)
    report_path = cv_dir / "cv5fold_report.txt"
    report_path.write_text(report_text, encoding="utf-8")
    tqdm.write(f"  5-fold CV done : {len(results)} layers  →  {cv_dir}")
    tqdm.write(f"  Report         → {report_path}")

    return results



def write_report(
    model_key: str,
    model_cfg: dict,
    out_dir: str,
    summary: list[dict],
    contrastive_summary: list[dict],
    all_pass_summary: list[dict],
    all_fail_summary: list[dict],
    probe_results: list[dict],
    train_ids: set | None = None,
    contrastive_test_summary: list[dict] | None = None,
    test_all_pass_summary: list[dict] | None    = None,
    test_all_fail_summary: list[dict] | None    = None,
) -> None:
    n_total  = len(summary)
    n_right  = sum(1 for s in summary if s["any_passed"])
    pct_pass = n_right / n_total * 100 if n_total else 0

    _ctest  = contrastive_test_summary  or []
    _tpass  = test_all_pass_summary     or []
    _tfail  = test_all_fail_summary     or []

    n_train_all       = len([s for s in summary if train_ids and s["task_id"] in train_ids]) \
                        if train_ids else 0
    n_contrastive     = len(contrastive_summary)
    n_all_pass        = len(all_pass_summary)
    n_all_fail        = len(all_fail_summary)
    n_te_contrastive  = len(_ctest)
    n_te_all_pass     = len(_tpass)
    n_te_all_fail     = len(_tfail)
    n_test_all        = n_te_contrastive + n_te_all_pass + n_te_all_fail

    passk_all   = _agg_passk(summary)
    passk_train = _agg_passk(contrastive_summary)
    passk_test  = _agg_passk(_ctest) if _ctest else {}

    json_path = Path(out_dir) / "summary.json"
    with open(json_path, "w") as f:
        json.dump({
            "model_key":   model_key,
            "display":     model_cfg["display"],
            "hf_id":       model_cfg["hf_id"],
            "labeling":    "1 = ALL test-cases pass; 0 = any failure",
            "activations": "token before EOS in generated sequence (decoder)",
            "mode":        "contrastive_both_splits — probe trained on mixed-verdict train problems, tested on mixed-verdict test problems",
            "n_total":     n_total,
            "n_right":     n_right,
            "n_wrong":     n_total - n_right,
            "pct_right":   round(pct_pass, 2),
            "split": {
                "n_train_all":    n_train_all,
                "n_test_all":     n_test_all,
            },
            "contrastive_filter": {
                "applied_to":           "both train and test splits",
                "train": {
                    "n_all":            n_train_all,
                    "n_contrastive":    n_contrastive,
                    "n_all_pass":       n_all_pass,
                    "n_all_fail":       n_all_fail,
                    "pct_contrastive":  round(
                        n_contrastive / n_train_all * 100, 2) if n_train_all else 0,
                    "contrastive_ids":  sorted(s["task_id"] for s in contrastive_summary),
                    "all_pass_ids":     sorted(s["task_id"] for s in all_pass_summary),
                    "all_fail_ids":     sorted(s["task_id"] for s in all_fail_summary),
                },
                "test": {
                    "n_all":            n_test_all,
                    "n_contrastive":    n_te_contrastive,
                    "n_all_pass":       n_te_all_pass,
                    "n_all_fail":       n_te_all_fail,
                    "pct_contrastive":  round(
                        n_te_contrastive / n_test_all * 100, 2) if n_test_all else 0,
                    "contrastive_ids":  sorted(s["task_id"] for s in _ctest),
                    "all_pass_ids":     sorted(s["task_id"] for s in _tpass),
                    "all_fail_ids":     sorted(s["task_id"] for s in _tfail),
                },
            },
            "pass_at_k": {
                "overall":           {f"pass@{k}": round(v, 4) for k, v in passk_all.items()},
                "contrastive_train": {f"pass@{k}": round(v, 4) for k, v in passk_train.items()},
                "test_all":          {f"pass@{k}": round(v, 4) for k, v in passk_test.items()},
            },
            "probing": probe_results,
        }, f, indent=2)
    tqdm.write(f"  JSON   → {json_path}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    except ImportError:
        tqdm.write("  openpyxl missing — Excel report skipped.")
        return

    PASS_F = PatternFill("solid", fgColor="C6EFCE")
    FAIL_F = PatternFill("solid", fgColor="FFC7CE")
    HEAD_F = PatternFill("solid", fgColor="D9E1F2")
    CONT_F = PatternFill("solid", fgColor="FFF2CC")
    BOLD   = Font(bold=True)
    CTR    = Alignment(horizontal="center", vertical="center")

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Overview"
    r = 1
    ws1.merge_cells(f"A{r}:G{r}")
    t = ws1.cell(r, 1, f"CONTRASTIVE PROBING — {model_cfg['display']}")
    t.font = Font(bold=True, size=15); t.fill = HEAD_F; t.alignment = CTR
    r += 1
    for label, val in [
        ("HF model ID",                          model_cfg["hf_id"]),
        ("Mode",                                 "Contrastive train / full test"),
        ("Timestamp",                            datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("N_RUNS",                               N_RUNS),
        ("Temperature",                          TEMPERATURE),
        ("Labeling",                             "1 = ALL asserts pass,  0 = any assert fails"),
        ("Problems (total generated)",           n_total),
        ("── TRAIN SET (80%) ──────────────────", ""),
        ("  Train problems (all)",               n_train_all),
        ("    → all-pass  (excluded from probe)", n_all_pass),
        ("    → all-fail  (excluded from probe)", n_all_fail),
        ("    → mixed / contrastive (probe train)",n_contrastive),
        ("    Contrastive % of train",           f"{n_contrastive/n_train_all*100:.1f}%" if n_train_all else "N/A"),
        ("── TEST SET  (20%) ──────────────────", ""),
        ("  Test problems (all)",               n_test_all),
        ("    → all-pass  (excluded from probe)", n_te_all_pass),
        ("    → all-fail  (excluded from probe)", n_te_all_fail),
        ("    → mixed / contrastive (probe test)", n_te_contrastive),
        ("    Contrastive % of test",           f"{n_te_contrastive/n_test_all*100:.1f}%" if n_test_all else "N/A"),
    ]:
        c1 = ws1.cell(r, 1, label); c1.font = BOLD
        c2 = ws1.cell(r, 2, str(val))
        if "mixed / contrastive" in label.lower():
            c1.fill = CONT_F; c2.fill = CONT_F
        r += 1

    r += 1
    ws1.merge_cells(f"A{r}:G{r}")
    hd = ws1.cell(r, 1, "pass@k  (unbiased estimator)")
    hd.font = BOLD; hd.fill = HEAD_F; hd.alignment = CTR
    r += 1
    ws1.cell(r, 1, "Split").font = BOLD
    for ci, k in enumerate(PASSK_VALUES, 2):
        ws1.cell(r, ci, f"pass@{k}").font = BOLD
    r += 1
    for split_label, pk_dict in [
        ("Overall (all 164)",                passk_all),
        ("Contrastive train (mixed-verdict)", passk_train),
        ("Test — all held-out problems",      passk_test),
    ]:
        ws1.cell(r, 1, split_label)
        for ci, k in enumerate(PASSK_VALUES, 2):
            v = pk_dict.get(k)
            ws1.cell(r, ci, round(v, 4) if v is not None else "N/A").alignment = CTR
        r += 1

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 50

    ws2 = wb.create_sheet("LinearProbing")
    if probe_results:
        hdr = [
            "layer_idx", "layer_name",
            "n_train_right", "n_train_wrong", "n_train_total",
            "cv_bal_acc_mean", "cv_bal_acc_std", "cv_acc_mean", "cv_auc_mean",
            "n_test_right", "n_test_wrong", "n_test_total",
            "test_bal_acc", "test_acc", "test_auc",
        ]
        for ci, h in enumerate(hdr, 1):
            c = ws2.cell(1, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
        ws2.freeze_panes = "A2"
        best_cv_val = max(r["cv_bal_acc_mean"] for r in probe_results)
        for ri, rec in enumerate(probe_results, 2):
            for ci, key in enumerate(hdr, 1):
                v = rec.get(key)
                cell = ws2.cell(ri, ci, v)
                cell.alignment = CTR
                if key == "cv_bal_acc_mean" and v == best_cv_val:
                    cell.fill = PatternFill("solid", fgColor="FFD700")
                elif key in ("test_bal_acc", "test_auc") and isinstance(v, float):
                    cell.fill = PASS_F if v >= 0.6 else FAIL_F

    xlsx_path = Path(out_dir) / "summary.xlsx"
    wb.save(xlsx_path)
    tqdm.write(f"  Excel  → {xlsx_path}")



def visualize_model(out_dir, contrastive_summary, probe_results, model_cfg,
                    contrastive_test_summary=None):
    if not MATPLOTLIB_OK:
        return

    plots_dir = Path(out_dir) / "plots"
    plots_dir.mkdir(exist_ok=True)
    display = model_cfg["display"]

    train_rows  = contrastive_summary
    test_rows   = contrastive_test_summary or []
    passk_train = _agg_passk(train_rows) if train_rows else {}
    passk_test  = _agg_passk(test_rows)  if test_rows  else {}
    COLORS = {"overall": "#4C72B0", "train": "#55A868", "test": "#C44E52"}

    fig, ax = plt.subplots(figsize=(8, 5))
    x, w = np.arange(len(PASSK_VALUES)), 0.3
    for offset, (lbl, pk, col) in enumerate([
        ("Contrastive Train 80%", passk_train, COLORS["train"]),
        ("Contrastive Test 20%",  passk_test,  COLORS["test"]),
    ]):
        vals  = [pk.get(k, 0) for k in PASSK_VALUES]
        rects = ax.bar(x + (offset - 0.5) * w, vals, w, label=lbl, color=col, alpha=0.85)
        for rect, v in zip(rects, vals):
            ax.text(rect.get_x() + rect.get_width() / 2,
                    rect.get_height() + 0.012,
                    f"{v:.3f}", ha="center", va="bottom", fontsize=8)
    ax.set_xticks(x)
    ax.set_xticklabels([f"pass@{k}" for k in PASSK_VALUES], fontsize=11)
    ax.set_ylim(0, 1.15)
    ax.set_ylabel("Score")
    ax.set_title(f"{display}  —  pass@k  (contrastive subset)")
    ax.legend(); ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(plots_dir / "passk_bar.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(8, 5))
    bins = np.arange(N_RUNS + 2) - 0.5
    if train_rows and test_rows:
        ax.hist([s["n_passed"] for s in train_rows], bins=bins,
                alpha=0.75, label="Train 80%", color=COLORS["train"], edgecolor="white")
        ax.hist([s["n_passed"] for s in test_rows], bins=bins,
                alpha=0.75, label="Test 20%",  color=COLORS["test"],  edgecolor="white")
        ax.legend()
    else:
        ax.hist([s["n_passed"] for s in contrastive_summary], bins=bins,
                alpha=0.85, color=COLORS["overall"], edgecolor="white")
    ax.set_xticks(range(N_RUNS + 1))
    ax.set_xlabel(f"Runs passed  (out of {N_RUNS})", fontsize=11)
    ax.set_ylabel("Problems")
    ax.set_title(f"{display}  —  Contrastive runs-passed distribution\n"
                 "(all-pass / all-fail excluded)")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(plots_dir / "n_passed_dist.png", dpi=150)
    plt.close(fig)

    if not probe_results:
        return

    layers  = [r["layer_idx"]       for r in probe_results]
    cv_mean = [r["cv_bal_acc_mean"]  for r in probe_results]
    cv_std  = [r["cv_bal_acc_std"]   for r in probe_results]
    cv_auc  = [r["cv_auc_mean"]      for r in probe_results]
    te_bal  = [r.get("test_bal_acc") if r.get("test_bal_acc") is not None else float("nan")
               for r in probe_results]
    te_auc  = [r.get("test_auc")     if r.get("test_auc")     is not None else float("nan")
               for r in probe_results]

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.plot(layers, cv_mean, "o-", color=COLORS["overall"], lw=2, label="CV bal-acc (train, mean)")
    ax.fill_between(layers,
                    [m - s for m, s in zip(cv_mean, cv_std)],
                    [m + s for m, s in zip(cv_mean, cv_std)],
                    alpha=0.18, color=COLORS["overall"])
    ax.plot(layers, te_bal, "s--", color=COLORS["test"], lw=2, label="Test bal-acc")
    ax.axhline(0.5, color="gray", lw=1, ls=":", label="Chance (0.5)")
    best_cv_i = int(np.argmax(cv_mean))
    valid_te  = [(i, v) for i, v in enumerate(te_bal) if not np.isnan(v)]
    if valid_te:
        best_te_i = max(valid_te, key=lambda t: t[1])[0]
        ax.axvline(layers[best_te_i], color=COLORS["test"], lw=1, ls="--", alpha=0.5)
        ax.text(layers[best_te_i] + 0.3, min(te_bal[best_te_i] - 0.04, 0.95),
                f"best test\nlayer {layers[best_te_i]}", fontsize=8, color=COLORS["test"])
    ax.axvline(layers[best_cv_i], color=COLORS["overall"], lw=1, ls="--", alpha=0.5)
    ax.text(layers[best_cv_i] + 0.3, min(cv_mean[best_cv_i] + 0.01, 0.98),
            f"best CV\nlayer {layers[best_cv_i]}", fontsize=8, color=COLORS["overall"])
    ax.set_xlabel("Layer index", fontsize=11)
    ax.set_ylabel("Balanced accuracy")
    ax.set_title(f"{display}  —  Contrastive probe accuracy vs layer")
    ax.set_ylim(max(0, float(np.nanmin(cv_mean + te_bal)) - 0.05), 1.05)
    ax.legend(); ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(plots_dir / "probe_accuracy_layer.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.plot(layers, cv_auc, "o-",  color=COLORS["overall"], lw=2, label="CV AUC (train)")
    ax.plot(layers, te_auc, "s--", color=COLORS["test"],    lw=2, label="Test AUC")
    ax.axhline(0.5, color="gray", lw=1, ls=":", label="Chance (0.5)")
    ax.set_xlabel("Layer index", fontsize=11)
    ax.set_ylabel("ROC-AUC")
    ax.set_title(f"{display}  —  Contrastive probe ROC-AUC vs layer")
    ax.set_ylim(max(0, float(np.nanmin(cv_auc + te_auc)) - 0.05), 1.05)
    ax.legend(); ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(plots_dir / "probe_auc_layer.png", dpi=150)
    plt.close(fig)

    valid_idx = [i for i, v in enumerate(te_bal) if not np.isnan(v)]
    if valid_idx:
        x_cv = [cv_mean[i] for i in valid_idx]
        y_te = [te_bal[i]  for i in valid_idx]
        l_te = [layers[i]  for i in valid_idx]
        fig, ax = plt.subplots(figsize=(6, 6))
        sc = ax.scatter(x_cv, y_te, c=l_te, cmap="viridis", s=70, zorder=3)
        plt.colorbar(sc, ax=ax, label="Layer index")
        lo = min(min(x_cv), min(y_te)) - 0.03
        hi = max(max(x_cv), max(y_te)) + 0.03
        ax.plot([lo, hi], [lo, hi], "k--", lw=1, alpha=0.4, label="y = x")
        ax.set_xlim(lo, hi); ax.set_ylim(lo, hi)
        ax.set_xlabel("CV bal-acc (train)", fontsize=11)
        ax.set_ylabel("Test bal-acc", fontsize=11)
        ax.set_title(f"{display}  —  Contrastive CV vs test per layer")
        ax.legend(fontsize=8); ax.grid(alpha=0.3)
        fig.tight_layout()
        fig.savefig(plots_dir / "cv_vs_test_scatter.png", dpi=150)
        plt.close(fig)

    tqdm.write(f"  Plots  → {plots_dir}")



def visualize_multi_model(model_results: list[dict], out_root: str) -> None:
    if not MATPLOTLIB_OK or len(model_results) < 2:
        return

    cmp_dir = Path(out_root) / "comparison"
    cmp_dir.mkdir(exist_ok=True)
    labels = [r["display"] for r in model_results]
    n      = len(labels)
    colors = [plt.cm.tab10(i / 10) for i in range(n)]

    fig, axes = plt.subplots(1, 2, figsize=(14, 6), sharey=True)
    for ax, (split_key, title) in zip(
        axes,
        [("passk_train", "Contrastive Train 80%"), ("passk_test", "Contrastive Test 20%")],
    ):
        x = np.arange(len(PASSK_VALUES))
        w = 0.8 / n
        for i, (res, col) in enumerate(zip(model_results, colors)):
            pk   = res[split_key]
            vals = [pk.get(k, 0) for k in PASSK_VALUES]
            bars = ax.bar(x + (i - n / 2 + 0.5) * w, vals, w,
                          label=res["display"], color=col, alpha=0.85)
            for bar, v in zip(bars, vals):
                if v > 0.02:
                    ax.text(bar.get_x() + bar.get_width() / 2,
                            bar.get_height() + 0.005,
                            f"{v:.2f}", ha="center", va="bottom", fontsize=6.5)
        ax.set_xticks(x)
        ax.set_xticklabels([f"pass@{k}" for k in PASSK_VALUES])
        ax.set_ylim(0, 1.05)
        ax.set_ylabel("Score")
        ax.set_title(f"pass@k  —  {title}")
        ax.legend(fontsize=8, loc="upper left")
        ax.grid(axis="y", alpha=0.3)
    fig.suptitle("Contrastive pass@k — all models", fontsize=13, fontweight="bold")
    fig.tight_layout()
    fig.savefig(cmp_dir / "models_passk.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(max(8, n * 1.8), 5))
    x = np.arange(n)
    w = 0.28
    ax.bar(x - w, [(r["n_train_all_pass"] + r["n_test_all_pass"])
                   / max(r["n_total"], 1) * 100 for r in model_results],
           w, label="All-pass (excluded)",  color="#55A868", alpha=0.8)
    ax.bar(x,     [(r["n_contrastive_train"] + r["n_contrastive_test"])
                   / max(r["n_total"], 1) * 100 for r in model_results],
           w, label="Contrastive (kept)",   color="#4C72B0", alpha=0.8)
    ax.bar(x + w, [(r["n_train_all_fail"] + r["n_test_all_fail"])
                   / max(r["n_total"], 1) * 100 for r in model_results],
           w, label="All-fail (excluded)",  color="#C44E52", alpha=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=20, ha="right", fontsize=10)
    ax.set_ylabel("% of problems")
    ax.set_title("Problem split by verdict type — all models", fontsize=12, fontweight="bold")
    ax.legend(); ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(cmp_dir / "contrastive_fraction.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(max(8, n * 1.6), 5))
    x, w = np.arange(n), 0.35
    ax.bar(x - w / 2, [r["best_cv_bal_acc"]  for r in model_results], w,
           label="Best CV bal-acc (train)", color="#4C72B0", alpha=0.85)
    ax.bar(x + w / 2, [r["best_test_bal_acc"] for r in model_results], w,
           label="Best test bal-acc",       color="#C44E52", alpha=0.85)
    ax.axhline(0.5, color="gray", lw=1, ls=":", label="Chance (0.5)")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=20, ha="right", fontsize=10)
    ax.set_ylim(0, 1.0)
    ax.set_ylabel("Balanced accuracy")
    ax.set_title("Best-layer contrastive probe accuracy  —  all models", fontsize=12, fontweight="bold")
    ax.legend(); ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(cmp_dir / "models_best_probe.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(7, 6))
    for res, col in zip(model_results, colors):
        p1  = res["passk_test"].get(1, 0)
        bte = res["best_test_bal_acc"]
        ax.scatter(p1, bte, s=130, color=col, zorder=3)
        ax.annotate(res["display"], (p1, bte),
                    textcoords="offset points", xytext=(7, 4), fontsize=9)
    ax.set_xlabel("pass@1  (contrastive test split)", fontsize=11)
    ax.set_ylabel("Best contrastive test probe bal-acc", fontsize=11)
    ax.set_title("Capability vs probe separability  (contrastive)", fontsize=11)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(cmp_dir / "capability_vs_probe.png", dpi=150)
    plt.close(fig)

    fig, axes = plt.subplots(1, 2, figsize=(16, 5))
    for ax, metric_key, title in zip(
        axes,
        ["cv_bal_acc_mean", "test_bal_acc"],
        ["CV bal-acc (train)", "Test bal-acc"],
    ):
        for res, col in zip(model_results, colors):
            pr = res["probe_results"]
            if not pr:
                continue
            lyr   = [r["layer_idx"] for r in pr]
            vals  = [r.get(metric_key) for r in pr]
            n_l   = max(lyr) if max(lyr) > 0 else 1
            x_rel = [l / n_l for l in lyr]
            y     = [v if v is not None else float("nan") for v in vals]
            ax.plot(x_rel, y, "o-", color=col, lw=1.8, ms=4,
                    label=res["display"], alpha=0.85)
        ax.axhline(0.5, color="gray", lw=1, ls=":", alpha=0.6)
        ax.set_xlabel("Relative layer depth (0=embed, 1=last)", fontsize=10)
        ax.set_ylabel("Balanced accuracy")
        ax.set_title(f"{title}  —  contrastive  (all models)")
        ax.legend(fontsize=8)
        ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(cmp_dir / "models_probe_by_layer.png", dpi=150)
    plt.close(fig)

    tqdm.write(f"\n  Comparison plots → {cmp_dir}")



def reorganize_output_dirs(out_dir: str, contrastive_ids: set) -> None:
    """
    After generation + probing are complete, reorganise all problem folders:

      {out_dir}/
        all/
          <task_safe>/   ← every problem (moved here)
          …
        contrastive/
          <task_safe>/   ← symlink → ../all/<task_safe>  (contrastive only)
          …

    Symlinks keep disk usage flat — 'contrastive/' is a lightweight index view.
    """
    _skip = {"probing", "plots", "all", "contrastive"}
    _safe_to_tid = {tid.replace("/", "_", 1): tid for tid in contrastive_ids}

    out_path = Path(out_dir)
    all_dir  = out_path / "all"
    cnt_dir  = out_path / "contrastive"
    all_dir.mkdir(exist_ok=True)
    cnt_dir.mkdir(exist_ok=True)

    moved = 0
    linked = 0

    for prob_dir in sorted(out_path.iterdir()):
        if not prob_dir.is_dir() or prob_dir.name in _skip:
            continue
        dest = all_dir / prob_dir.name
        shutil.move(str(prob_dir), str(dest))
        moved += 1

        if prob_dir.name in _safe_to_tid:
            link = cnt_dir / prob_dir.name
            link.symlink_to(Path("..") / "all" / prob_dir.name)
            linked += 1

    tqdm.write(
        f"  Reorganised: {moved} folders → all/  |  "
        f"{linked} symlinks → contrastive/"
    )



def _find_existing_run(out_root: str, model_key: str) -> str | None:
    """
    Scan out_root for a previously completed run for model_key.
    A run is considered complete if its directory contains split.json
    (written after generation + contrastive filtering are done).
    Returns the path of the most recent such directory, or None.
    """
    root = Path(out_root)
    if not root.exists():
        return None
    candidates = [
        d for d in root.iterdir()
        if d.is_dir()
        and d.name.startswith(f"{model_key}_")
        and (d / "split.json").exists()
    ]
    if not candidates:
        return None
    return str(sorted(candidates)[-1])



def recompute_split_from_split_json(out_dir: str) -> dict:
    """
    Re-split contrastive IDs from an existing split.json using the stratified
    70/30 approach (split only contrastive problems, not all problems).

    Reads split.json, pools all contrastive IDs, applies make_split to those
    IDs only, and writes the updated split.json back in-place.

    Returns the updated split dict.
    """
    split_path = Path(out_dir) / "split.json"
    with open(split_path) as f:
        sp = json.load(f)

    old_train = set(sp["train"]["contrastive_ids"])
    old_test  = set(sp["test"]["contrastive_ids"])
    all_cont  = sorted(old_train | old_test)

    cont_indexed = [(i, {"task_id": tid}) for i, tid in enumerate(all_cont)]
    train_list, test_list, new_train_ids, new_test_ids = make_split(cont_indexed)

    sp["train"]["contrastive_ids"] = sorted(new_train_ids)
    sp["train"]["n_contrastive"]   = len(new_train_ids)
    sp["test"]["contrastive_ids"]  = sorted(new_test_ids)
    sp["test"]["n_contrastive"]    = len(new_test_ids)
    if all_cont:
        sp["train"]["pct_contrastive"] = round(len(new_train_ids) / len(all_cont) * 100, 2)
        sp["test"]["pct_contrastive"]  = round(len(new_test_ids)  / len(all_cont) * 100, 2)
    sp["stratified_contrastive_split"] = True

    with open(split_path, "w") as f:
        json.dump(sp, f, indent=2)

    tqdm.write(
        f"  [split recomputed]  train_cntr={len(new_train_ids)}  "
        f"test_cntr={len(new_test_ids)}  total={len(all_cont)}"
    )
    return sp



def select_dataset(args_dataset: str | None = None,
                   apps_difficulty: str = "introductory",
                   apps_max: int = 500) -> tuple[str, list[dict]]:
    """Always loads MBPP+ (the only dataset in the CODE_LLM pipeline).
    args_dataset is accepted for compatibility but must be 'mbppplus' or None."""
    if args_dataset is not None and args_dataset.lower() != "mbppplus":
        raise ValueError(
            f"CODE_LLM pipeline only supports 'mbppplus', got '{args_dataset}'."
        )
    problems = load_mbppplus()
    return "mbppplus", problems



def main() -> None:
    parser = argparse.ArgumentParser(
        description="Qwen2.5-Coder — Contrastive layer-wise activation extraction (MBPP+)"
    )
    parser.add_argument("--dataset", default=None,
                        choices=["mbppplus"],
                        help="Dataset (only 'mbppplus' supported)")
    grp = parser.add_mutually_exclusive_group()
    grp.add_argument("--all",   action="store_true")
    grp.add_argument("--index", type=int, default=None)
    grp.add_argument("--id",    type=str, default=None)
    parser.add_argument("--start",  type=int, default=0)
    parser.add_argument("--end",    type=int, default=None)
    parser.add_argument("--models", nargs="+", default=None,
                        choices=list(MODEL_REGISTRY.keys()))
    parser.add_argument("--skip-probing", action="store_true")
    parser.add_argument("--recompute-split", action="store_true",
                        help="Re-split existing split.json files to stratified 70/30 "
                             "contrastive-only split and exit (no generation or probing).")
    parser.add_argument("--list", action="store_true")
    args = parser.parse_args()
    global OUT_ROOT

    if args.recompute_split:
        ds_key_fix = args.dataset.lower() if args.dataset else "mbppplus"
        OUT_ROOT   = _MBPPPLUS_ROOT
        model_keys_fix = args.models or list(MODEL_REGISTRY.keys())
        print(f"\n  Recomputing stratified 70/30 splits for [{ds_key_fix}] …")
        fixed = 0
        for mk in model_keys_fix:
            existing = _find_existing_run(OUT_ROOT, mk)
            if not existing:
                print(f"  {MODEL_REGISTRY[mk]['display']:25}  no existing run — skipped")
                continue
            sp_path = Path(existing) / "split.json"
            if not sp_path.exists():
                print(f"  {MODEL_REGISTRY[mk]['display']:25}  no split.json — skipped")
                continue
            with open(sp_path) as f:
                sp = json.load(f)
            if sp.get("stratified_contrastive_split"):
                tr   = len(sp["train"]["contrastive_ids"])
                te   = len(sp["test"]["contrastive_ids"])
                tot  = tr + te
                diff = sp.get("apps_difficulty", "")
                diff_str = f"  [{diff}]" if diff else ""
                pct  = f"{te/tot*100:.0f}% test" if tot else "?"
                print(f"  {MODEL_REGISTRY[mk]['display']:25}  already stratified{diff_str}  train={tr}  test={te}  ({pct})")
                continue
            sp = recompute_split_from_split_json(existing)
            tr  = len(sp["train"]["contrastive_ids"])
            te  = len(sp["test"]["contrastive_ids"])
            tot = tr + te
            pct = f"{te/tot*100:.0f}% test" if tot else "?"
            print(f"  {MODEL_REGISTRY[mk]['display']:25}  fixed  train={tr}  test={te}  ({pct})")
            fixed += 1
        print(f"\n  Done — {fixed} split(s) recomputed.")
        return

    ds_key, problems = select_dataset(
        args_dataset    = args.dataset,
        apps_difficulty = args.apps_difficulty,
        apps_max        = args.apps_max,
    )

    OUT_ROOT = _MBPPPLUS_ROOT
    os.makedirs(OUT_ROOT, exist_ok=True)

    if args.list:
        print(f"\n{'IDX':>4}  {'task_id':>20}  entry_point / mode")
        print("-" * 55)
        for i, p in enumerate(problems):
            ep = p.get("entry_point") or p.get("execution_mode", "?")
            print(f"{i:>4}  {p['task_id']:>20}  {ep}")
        return

    if not args.all and args.index is None and args.id is None:
        args.all = True

    if args.all:
        end      = args.end if args.end is not None else len(problems)
        selected = [(i, problems[i]) for i in range(args.start, min(end, len(problems)))]
    elif args.index is not None:
        selected = [(args.index, problems[args.index])]
    else:
        p = next((q for q in problems if q["task_id"] == args.id), None)
        if p is None:
            parser.error(f"task_id '{args.id}' not found.")
        idx      = next(i for i, q in enumerate(problems) if q["task_id"] == p["task_id"])
        selected = [(idx, p)]

    model_keys = args.models or list(MODEL_REGISTRY.keys())
    device     = "cuda" if torch.cuda.is_available() else "cpu"
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(OUT_ROOT, exist_ok=True)

    print(f"\n{'='*66}")
    print(f"  Qwen2.5-Coder  —  Contrastive Layer-wise Activation Extraction")
    print(f"  Dataset   : {ds_key}  ({len(selected)} problems selected)")
    print(f"  Output    : {OUT_ROOT}/")
    print(f"  Models    : {len(model_keys)}  |  N_RUNS : {N_RUNS}  |  Temp : {TEMPERATURE}")
    print(f"  Contrastive filter : 0 < n_passed < {N_RUNS}  (mixed-verdict only)")
    print(f"  Split     : {int(TRAIN_RATIO*100)}/{100-int(TRAIN_RATIO*100)}  "
          f"seed={SPLIT_SEED}")
    print(f"  Device    : {device}")
    print(f"{'='*66}\n")

    all_model_results: list[dict] = []

    model_pbar = tqdm(model_keys, desc="Models", unit="model", position=0, leave=True)
    for model_key in model_pbar:
        model_cfg = MODEL_REGISTRY[model_key]
        model_pbar.set_description(f"Model: {model_cfg['display']}")

        existing = _find_existing_run(OUT_ROOT, model_key)
        if existing:
            out_dir = existing
            tqdm.write(f"\n{'#'*66}")
            tqdm.write(f"  {model_cfg['display']}  —  RESUMING existing run")
            tqdm.write(f"  Directory : {out_dir}")
            tqdm.write(f"{'#'*66}")

            try:
                with open(Path(out_dir) / "split.json") as f:
                    sp = json.load(f)
                contrastive_train_ids = set(sp["train"]["contrastive_ids"])
                contrastive_test_ids  = set(sp["test"]["contrastive_ids"])
            except Exception as exc:
                tqdm.write(f"  ERROR reading split.json: {exc}  — skipping {model_key}")
                continue

            all_contrastive_ids = contrastive_train_ids | contrastive_test_ids

            if not sp.get("stratified_contrastive_split"):
                tqdm.write(
                    f"  [WARN] split.json uses old all-problems split "
                    f"(train={len(contrastive_train_ids)}, test={len(contrastive_test_ids)}) "
                    f"— recomputing stratified 70/30 contrastive split …"
                )
                sp = recompute_split_from_split_json(out_dir)
                contrastive_train_ids = set(sp["train"]["contrastive_ids"])
                contrastive_test_ids  = set(sp["test"]["contrastive_ids"])
                all_contrastive_ids   = contrastive_train_ids | contrastive_test_ids

            tqdm.write(
                f"  Loaded split.json: train_cntr={len(contrastive_train_ids)}  "
                f"test_cntr={len(contrastive_test_ids)}  "
                f"total={len(all_contrastive_ids)}"
            )

            if not args.skip_probing and contrastive_train_ids:
                probe_json = Path(out_dir) / "probing" / "probing_results.json"
                if probe_json.exists():
                    tqdm.write(f"  Regular probing already done — skipping.")
                else:
                    tqdm.write(
                        f"  Regular probing not found — running now  "
                        f"(train: {len(contrastive_train_ids)}  "
                        f"test: {len(contrastive_test_ids)})"
                    )
                    probe_results = run_linear_probing(
                        out_dir, contrastive_train_ids, contrastive_test_ids
                    )
                    if probe_results:
                        best = max(probe_results, key=lambda r: r["cv_bal_acc_mean"])
                        te   = best.get("test_bal_acc")
                        tqdm.write(
                            f"  Best layer : {best['layer_name']}  "
                            f"CV bal_acc={best['cv_bal_acc_mean']:.3f}±{best['cv_bal_acc_std']:.3f}"
                            + (f"  test bal_acc={te:.3f}" if te is not None else "")
                        )

                if all_contrastive_ids:
                    cv5_results = run_5fold_cv_all(
                        out_dir, all_contrastive_ids, model_cfg["display"]
                    )
                    if cv5_results:
                        best5 = max(cv5_results, key=lambda r: r["cv_bal_acc_mean"])
                        tqdm.write(
                            f"  5-fold CV best : {best5['layer_name']}  "
                            f"bal_acc={best5['cv_bal_acc_mean']:.3f}±{best5['cv_bal_acc_std']:.3f}  "
                            f"AUC={best5['cv_auc_mean']:.3f}"
                        )

            tqdm.write(f"  Resumed  : {model_cfg['display']}\n")
            continue

        out_dir = os.path.join(OUT_ROOT, f"{model_key}_{ts}")
        os.makedirs(out_dir, exist_ok=True)

        tqdm.write(f"\n{'#'*66}")
        tqdm.write(f"  {model_cfg['display']}  ({model_cfg['hf_id']})")
        tqdm.write(f"  Output → {out_dir}")
        tqdm.write(f"{'#'*66}")

        summary, _ = run_analysis(
            model_key, model_cfg, selected, device, out_dir
        )

        n_total = len(summary)

        train_list, test_list, train_ids_all, test_ids_all = make_split(selected)

        all_cont_indexed, _, _, _ = filter_contrastive(summary, selected)
        (
            train_cont_list, test_cont_list,
            contrastive_train_ids, contrastive_test_ids,
        ) = make_split(all_cont_indexed)

        contrastive_train_summary = [s for s in summary if s["task_id"] in contrastive_train_ids]
        contrastive_test_summary  = [s for s in summary if s["task_id"] in contrastive_test_ids]
        n_tr_contrastive = len(contrastive_train_summary)
        n_te_contrastive = len(contrastive_test_summary)

        _, _, train_all_pass_summary, train_all_fail_summary = filter_contrastive(summary, train_list)
        _, _, test_all_pass_summary,  test_all_fail_summary  = filter_contrastive(summary, test_list)

        tqdm.write(
            f"  70/30 contrastive split: train {n_tr_contrastive}  test {n_te_contrastive}"
            f"  (total contrastive: {len(all_cont_indexed)})"
        )

        report_contrastive_stats(
            [s for s in summary if s["task_id"] in train_ids_all],
            contrastive_train_summary,
            train_all_pass_summary,
            train_all_fail_summary,
            [s for s in summary if s["task_id"] in test_ids_all],
            contrastive_test_summary,
            test_all_pass_summary,
            test_all_fail_summary,
            out_dir, model_cfg["display"],
        )

        if n_tr_contrastive < 4:
            tqdm.write(
                f"  WARNING: only {n_tr_contrastive} contrastive train problems — "
                "probing will be unreliable."
            )

        with open(os.path.join(out_dir, "split.json"), "w") as f:
            _split_meta = {
                "mode":                          "contrastive_both_splits",
                "train_ratio":                   TRAIN_RATIO,
                "split_seed":                    SPLIT_SEED,
                "n_total":                       n_total,
                "stratified_contrastive_split":  True,
            }
            if ds_key == "apps":
                _split_meta["apps_difficulty"] = args.apps_difficulty
            json.dump({**_split_meta, "train": {
                    "n_all":                 len(train_list),
                    "n_contrastive":         n_tr_contrastive,
                    "n_all_pass":            len(train_all_pass_summary),
                    "n_all_fail":            len(train_all_fail_summary),
                    "pct_contrastive":       round(n_tr_contrastive / len(all_cont_indexed) * 100, 2)
                                             if all_cont_indexed else 0,
                    "contrastive_ids":       sorted(contrastive_train_ids),
                    "all_pass_ids":          sorted(s["task_id"] for s in train_all_pass_summary),
                    "all_fail_ids":          sorted(s["task_id"] for s in train_all_fail_summary),
                },
                "test": {
                    "n_all":                 len(test_list),
                    "n_contrastive":         n_te_contrastive,
                    "n_all_pass":            len(test_all_pass_summary),
                    "n_all_fail":            len(test_all_fail_summary),
                    "pct_contrastive":       round(n_te_contrastive / len(all_cont_indexed) * 100, 2)
                                             if all_cont_indexed else 0,
                    "contrastive_ids":       sorted(contrastive_test_ids),
                    "all_pass_ids":          sorted(s["task_id"] for s in test_all_pass_summary),
                    "all_fail_ids":          sorted(s["task_id"] for s in test_all_fail_summary),
                },
            }, f, indent=2)

        probe_results: list[dict] = []
        if not args.skip_probing and contrastive_train_ids:
            tqdm.write(
                f"  Running contrastive probing …"
                f"  train: {n_tr_contrastive}  |  test: {n_te_contrastive}"
            )
            probe_results = run_linear_probing(
                out_dir, contrastive_train_ids, contrastive_test_ids
            )
            if probe_results:
                best = max(probe_results, key=lambda r: r["cv_bal_acc_mean"])
                te   = best.get("test_bal_acc")
                tqdm.write(
                    f"  Best layer : {best['layer_name']}  "
                    f"CV bal_acc={best['cv_bal_acc_mean']:.3f}±{best['cv_bal_acc_std']:.3f}"
                    + (f"  test bal_acc={te:.3f}" if te is not None else "")
                )

        all_contrastive_ids      = contrastive_train_ids | contrastive_test_ids
        all_contrastive_summary  = contrastive_train_summary + contrastive_test_summary
        cv5_results: list[dict]  = []
        if not args.skip_probing and all_contrastive_ids:
            cv5_results = run_5fold_cv_all(
                out_dir,
                all_contrastive_ids,
                model_cfg["display"],
            )
            if cv5_results:
                best5 = max(cv5_results, key=lambda r: r["cv_bal_acc_mean"])
                tqdm.write(
                    f"  5-fold CV best : {best5['layer_name']}  "
                    f"bal_acc={best5['cv_bal_acc_mean']:.3f}±{best5['cv_bal_acc_std']:.3f}  "
                    f"AUC={best5['cv_auc_mean']:.3f}"
                )

        write_report(
            model_key, model_cfg, out_dir,
            summary,
            contrastive_train_summary, train_all_pass_summary, train_all_fail_summary,
            probe_results,
            train_ids=contrastive_train_ids,
            contrastive_test_summary=contrastive_test_summary,
            test_all_pass_summary=test_all_pass_summary,
            test_all_fail_summary=test_all_fail_summary,
        )
        visualize_model(
            out_dir, contrastive_train_summary, probe_results, model_cfg,
            contrastive_test_summary=contrastive_test_summary,
        )

        reorganize_output_dirs(out_dir, all_contrastive_ids)

        best_cv = max((r["cv_bal_acc_mean"]   for r in probe_results), default=0.0)
        best_te = max((r["test_bal_acc"] or 0 for r in probe_results), default=0.0)

        passk_train_d = _agg_passk(contrastive_train_summary)
        passk_test_d  = _agg_passk(contrastive_test_summary)

        all_model_results.append({
            "display":                model_cfg["display"],
            "n_total":                n_total,
            "n_contrastive_train":    n_tr_contrastive,
            "n_contrastive_test":     n_te_contrastive,
            "n_train_all_pass":       len(train_all_pass_summary),
            "n_train_all_fail":       len(train_all_fail_summary),
            "n_test_all_pass":        len(test_all_pass_summary),
            "n_test_all_fail":        len(test_all_fail_summary),
            "passk_train":            passk_train_d,
            "passk_test":             passk_test_d,
            "best_cv_bal_acc":        best_cv,
            "best_test_bal_acc":      best_te,
            "probe_results":          probe_results,
        })

        tqdm.write(f"  Finished : {model_cfg['display']}\n")

    visualize_multi_model(all_model_results, OUT_ROOT)


if __name__ == "__main__":
    main()
