"""
CODE_LLM steering.py  —  Step 4: Activation steering using directions from probing.py
======================================================================================
Uses steering directions computed by probing.py (Step 2) to steer Qwen models
on MBPP+ contrastive test problems.

Reads from:   ../mbppplus/<model_key>_<ts>/probing/  (directions, top_layers.json)
Writes to:    ../mbppplus/<model_key>_<ts>/steering/

Output structure (inside model_dir/steering/):
  pipeline2_results.json    — full results
  steering_summary.csv
  steering_report.txt
  steering_report.xlsx
  plots/

Usage
-----
  # Interactive (auto-discovers probing-complete runs)
  python steering.py

  # CLI single model
  python steering.py \\
      --model       qwen-coder-1.5b-instruct \\
      --acts-dir    /path/to/mbppplus/qwen-coder-1.5b-instruct_xxx \\
      --results-dir /path/to/mbppplus/qwen-coder-1.5b-instruct_xxx \\
      [--alphas-fwd 0.5 1.0 2.0 5.0 10.0] \\
      [--device cuda]
"""

import argparse
import csv
import gc
import json
import math
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import torch
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from tqdm import tqdm

import importlib.util as _ilu

_GEN_PATH = Path(__file__).parent / "contrastive_set_gen.py"
_spec     = _ilu.spec_from_file_location("_gen", _GEN_PATH)
_gen      = _ilu.module_from_spec(_spec)
_spec.loader.exec_module(_gen)

MODEL_REGISTRY                    = _gen.MODEL_REGISTRY
load_model                        = _gen.load_model
unload_model                      = _gen.unload_model
build_prompt                      = _gen.build_prompt
evaluate_all_cases                = _gen.evaluate_all_cases
truncate_to_function_continuation = _gen.truncate_to_function_continuation
extract_full_program              = _gen.extract_full_program
select_dataset                    = _gen.select_dataset
pass_at_k                         = _gen.pass_at_k

_DS_DISPLAY = {
    "mbppplus": "MBPP+  (~378 problems, more robust tests — evalplus)",
}


def _infer_model_key(dir_name: str) -> str | None:
    for mk in MODEL_REGISTRY:
        if dir_name.startswith(mk + "_") or dir_name == mk:
            return mk
    return None


_MODEL_KEY_ORDER = [
    "qwen-coder-1.5b-instruct",
    "qwen-coder-7b-instruct",
]



def _get_decoder_layer(model, layer_idx: int):
    """Return the decoder layer module at the given index."""
    for attr in ("model", "transformer", "gpt_neox", "language_model"):
        sub = getattr(model, attr, None)
        if sub is not None:
            for lattr in ("layers", "h", "blocks"):
                obj = getattr(sub, lattr, None)
                if isinstance(obj, torch.nn.ModuleList):
                    return obj[layer_idx]
    raise RuntimeError(
        f"Cannot find decoder layer {layer_idx} in model."
        "  Supported containers: model/transformer/gpt_neox/language_model → layers/h/blocks"
    )


class SteeringHook:
    """Context manager that adds α·direction to hidden states at one decoder layer."""

    def __init__(self, model, layer_idx: int, direction: torch.Tensor, alpha: float):
        self._layer     = _get_decoder_layer(model, layer_idx)
        self._direction = direction
        self._alpha     = alpha
        self._handle    = None

    def __enter__(self):
        alpha     = self._alpha
        direction = self._direction

        def _hook(module, input, output):
            hidden = output[0] if isinstance(output, tuple) else output
            hidden = hidden + alpha * direction
            return (hidden,) + output[1:] if isinstance(output, tuple) else hidden

        self._handle = self._layer.register_forward_hook(_hook)
        return self

    def __exit__(self, *_):
        if self._handle is not None:
            self._handle.remove()
            self._handle = None


def compute_baseline(acts_dir: str, test_ids: set) -> dict:
    """Count stored right/wrong runs per problem from the H5 directory structure.

    Returns {task_id: {"n_right", "n_wrong", "n_total", "pass@1"}}.
    """
    root = Path(acts_dir)
    if (root / "all").is_dir():
        root = root / "all"

    results: dict = {}
    for tid in test_ids:
        safe     = tid.replace("/", "_")
        prob_dir = root / safe
        if not prob_dir.is_dir():
            continue
        r_dir = prob_dir / "right"
        w_dir = prob_dir / "wrong"
        n_r   = len([d for d in r_dir.iterdir() if d.is_dir()]) if r_dir.is_dir() else 0
        n_w   = len([d for d in w_dir.iterdir() if d.is_dir()]) if w_dir.is_dir() else 0
        n_t   = n_r + n_w
        results[tid] = {
            "n_right": n_r,
            "n_wrong": n_w,
            "n_total": n_t,
            "pass@1":  pass_at_k(n_t, n_r, 1) if n_t > 0 else 0.0,
        }
    return results


def _agg_passk(per_prob: dict, k: int = 1) -> float:
    """Average pass@k across all problems that have a valid n_total > 0."""
    key  = f"pass@{k}"
    vals = [v[key] for v in per_prob.values()
            if key in v and v.get("n_total", 0) > 0]
    return float(np.mean(vals)) if vals else 0.0



DEFAULT_ALPHAS_FWD   = [0.5, 1.0, 2.0, 5.0, 10.0]
DEFAULT_ALPHAS_BWD   = [0.5, 1.0, 2.0, 5.0, 10.0]
N_STEER_RUNS         = 5
MAX_NEW_TOKENS       = 512
DEFAULT_RESULTS_ROOT = str(Path(__file__).parent.parent / "mbppplus")



def _find_p2_runs_with_probing(results_root: Path, dataset: str) -> list[tuple[str, Path]]:
    """
    Return [(model_key, results_dir), ...] for all pipeline2 results dirs
    that have completed probing (contain probing/probe_analysis.json).
    Only the most-recent run per model_key is kept.
    """
    ds_dir = results_root / dataset
    if not ds_dir.exists():
        return []
    best: dict[str, Path] = {}
    for d in sorted(ds_dir.iterdir()):
        if not d.is_dir():
            continue
        if not (d / "probing" / "probe_analysis.json").exists():
            continue
        mk = _infer_model_key(d.name)
        if mk is not None:
            best[mk] = d
    ordered  = [(mk, best[mk]) for mk in _MODEL_KEY_ORDER if mk in best]
    ordered += [(mk, best[mk]) for mk in best if mk not in _MODEL_KEY_ORDER]
    return ordered


def _find_p2_run(results_root: Path, dataset: str, model_key: str,
                 require_file: str | None = None) -> Path | None:
    ds_dir = results_root / dataset
    if not ds_dir.exists():
        return None
    candidates = [
        d for d in sorted(ds_dir.iterdir())
        if d.is_dir() and d.name.startswith(model_key + "_")
    ]
    if require_file:
        candidates = [d for d in candidates if (d / require_file).exists()]
    return candidates[-1] if candidates else None



def load_directions(results_dir: Path) -> dict[str, np.ndarray]:
    dir_path = results_dir / "probing" / "directions"
    if not dir_path.exists():
        raise FileNotFoundError(
            f"Directions not found: {dir_path}\nRun probing.py first."
        )
    directions: dict[str, np.ndarray] = {}
    for npy in sorted(dir_path.glob("layer_*.npy")):
        directions[npy.stem] = np.load(npy).astype(np.float32)
    if not directions:
        raise RuntimeError(f"No .npy files in {dir_path}")
    return directions


def load_top_layers(results_dir: Path) -> list[dict]:
    path = results_dir / "probing" / "top_layers.json"
    if not path.exists():
        raise FileNotFoundError(f"top_layers.json not found: {path}")
    with open(path) as f:
        return json.load(f)


def load_probe_analysis(results_dir: Path) -> dict:
    path = results_dir / "probing" / "probe_analysis.json"
    if not path.exists():
        raise FileNotFoundError(f"probe_analysis.json not found: {path}")
    with open(path) as f:
        return json.load(f)



def _apps_numeric_id(task_id: str) -> int | None:
    if task_id.startswith("apps/"):
        try:
            return int(task_id[5:])
        except ValueError:
            pass
    return None


def load_problems_with_apps_fix(
    ds_key: str,
    test_ids: set[str],
    apps_difficulty: str = "introductory",
    apps_max: int = 500,
) -> dict[str, dict]:
    print(f"  Loading {ds_key} problems …")
    _, all_problems = select_dataset(
        args_dataset=ds_key,
        apps_difficulty=apps_difficulty,
        apps_max=apps_max,
    )
    problem_map = {p["task_id"]: p for p in all_problems}

    if ds_key != "apps":
        return problem_map

    apps_test_ids = [tid for tid in test_ids if tid.startswith("apps/")]
    if not apps_test_ids:
        return problem_map

    numeric_ids = [n for tid in apps_test_ids if (n := _apps_numeric_id(tid)) is not None]
    max_numeric = max(numeric_ids, default=0)
    missing     = {tid for tid in apps_test_ids if tid not in problem_map}

    if missing or max_numeric >= apps_max:
        print(
            f"  [APPS fix] {len(missing)} test IDs missing"
            f" (max numeric id = {max_numeric} ≥ apps_max = {apps_max}).\n"
            f"  Reloading APPS without limit (difficulty='{apps_difficulty}') …"
        )
        for p in load_apps(difficulty=apps_difficulty, max_problems=None):
            problem_map[p["task_id"]] = p

    still_missing = {tid for tid in apps_test_ids if tid not in problem_map}
    if still_missing and apps_difficulty != "all":
        print(
            f"  [APPS fix] Still {len(still_missing)} missing — "
            "reloading difficulty='all' …"
        )
        for p in load_apps(difficulty="all", max_problems=None):
            problem_map[p["task_id"]] = p

    return problem_map



def generate_steered_p2(
    model,
    tokenizer,
    model_cfg:    dict,
    test_problems: dict,
    direction_np: np.ndarray,
    layer_idx:    int,
    alpha:        float,
    n_runs:       int | dict,
    device:       str,
) -> dict:
    direction = torch.from_numpy(direction_np).to(device)
    results: dict = {}

    pbar = tqdm(
        sorted(test_problems.keys()),
        desc=f"  α={alpha:+.1f}  layer={layer_idx:02d}",
        leave=False, unit="prob",
    )
    for tid in pbar:
        problem    = test_problems[tid]
        runs_this  = n_runs[tid] if isinstance(n_runs, dict) else n_runs
        if runs_this == 0:
            results[tid] = {"n_right": 0, "n_wrong": 0, "n_total": 0, "pass@1": 0.0}
            continue

        prompt = build_prompt(problem, model_cfg, tokenizer)
        inputs = tokenizer(prompt, return_tensors="pt").to(device)

        n_right = 0
        with SteeringHook(model, layer_idx, direction, alpha):
            for _ in range(runs_this):
                with torch.no_grad():
                    out_ids = model.generate(
                        **inputs,
                        max_new_tokens=MAX_NEW_TOKENS,
                        do_sample=True,
                        temperature=0.8,
                        top_p=0.95,
                        pad_token_id=tokenizer.eos_token_id,
                    )
                raw = tokenizer.decode(
                    out_ids[0][inputs["input_ids"].shape[1]:],
                    skip_special_tokens=True,
                )
                if problem.get("execution_mode") == "stdin_stdout":
                    code = extract_full_program(raw)
                else:
                    compl = truncate_to_function_continuation(raw, problem["prompt"])
                    code  = problem["prompt"] + compl
                try:
                    passed = evaluate_all_cases(code, problem)
                except Exception:
                    passed = False
                if passed:
                    n_right += 1

        results[tid] = {
            "n_right": n_right,
            "n_wrong": runs_this - n_right,
            "n_total": runs_this,
            "pass@1":  pass_at_k(runs_this, n_right, 1),
        }
    return results



def run_layer_experiment_p2(
    model,
    tokenizer,
    model_cfg:           dict,
    test_problems:       dict,
    direction_np:        np.ndarray,
    layer_idx:           int,
    layer_name:          str,
    auroc_cont:          float | None,
    run_counts:          dict,
    alphas_fwd:          list[float],
    alphas_bwd:          list[float],
    device:              str,
) -> dict:
    base_p1 = _agg_passk(
        {tid: v for tid, v in run_counts.items() if tid in test_problems}, 1
    )

    exp1_runs = {tid: run_counts[tid]["n_wrong"] for tid in test_problems if tid in run_counts}
    exp2_runs = {tid: run_counts[tid]["n_right"]  for tid in test_problems if tid in run_counts}

    n_exp1_total = sum(exp1_runs.values())
    n_exp2_total = sum(exp2_runs.values())
    auroc_str    = f"{auroc_cont:.3f}" if auroc_cont is not None else "N/A"

    tqdm.write(f"\n  ── {layer_name}  (AUROC_cont={auroc_str}) ──")
    tqdm.write(f"  Baseline pass@1 = {base_p1*100:.1f}%  (from stored runs, n={len(test_problems)} problems)")
    tqdm.write(f"  Exp1: {n_exp1_total} wrong runs to steer  |  Exp2: {n_exp2_total} right runs to steer")

    tqdm.write("  Experiment 1 — Wrong → Right  (α > 0)")
    exp1: list[dict] = []
    for alpha in alphas_fwd:
        per_prob = generate_steered_p2(
            model, tokenizer, model_cfg,
            test_problems, direction_np,
            layer_idx, alpha, exp1_runs, device,
        )
        p1  = _agg_passk(per_prob, 1)
        dp1 = (p1 - base_p1) * 100
        exp1.append({
            "alpha":       alpha,
            "pass@1":      p1,
            "delta_p1":    dp1,
            "per_problem": per_prob,
        })
        tqdm.write(f"    α={alpha:+.1f}  p@1={p1*100:.1f}%  Δp@1={dp1:+.1f}%")

    tqdm.write("  Experiment 2 — Right → Wrong  (α < 0)")
    exp2: list[dict] = []
    for alpha in alphas_bwd:
        per_prob = generate_steered_p2(
            model, tokenizer, model_cfg,
            test_problems, direction_np,
            layer_idx, alpha, exp2_runs, device,
        )
        p1  = _agg_passk(per_prob, 1)
        dp1 = (p1 - base_p1) * 100
        exp2.append({
            "alpha":       alpha,
            "pass@1":      p1,
            "delta_p1":    dp1,
            "per_problem": per_prob,
        })
        tqdm.write(f"    α={alpha:.1f}  p@1={p1*100:.1f}%  Δp@1={dp1:+.1f}%")

    best_fwd_p1 = max((r["delta_p1"] for r in exp1), default=0.0)
    best_bwd_p1 = min((r["delta_p1"] for r in exp2), default=0.0)
    tqdm.write(f"  → Best Fwd Δp@1={best_fwd_p1:+.1f}%  Best Bwd Δp@1={best_bwd_p1:+.1f}%")

    return {
        "layer_idx":        layer_idx,
        "layer_name":       layer_name,
        "auroc_cont":       auroc_cont,
        "baseline_p1":      base_p1,
        "n_wrong_total":    n_exp1_total,
        "n_right_total":    n_exp2_total,
        "n_test":           len(test_problems),
        "exp1":             exp1,
        "exp2":             exp2,
        "best_forward_p1":  best_fwd_p1,
        "best_backward_p1": best_bwd_p1,
    }


def _strip_per_problem(rows: list[dict]) -> list[dict]:
    return [{k: v for k, v in r.items() if k != "per_problem"} for r in rows]



def write_steering_report_txt(
    steer_dir:       Path,
    model_display:   str,
    dataset_display: str,
    layer_results:   list[dict],
) -> None:
    SEP  = "=" * 76
    SEP2 = "-" * 76
    lines: list[str] = []

    lines += [
        SEP,
        f"  PIPELINE-2 STEERING REPORT — {model_display}",
        f"  Dataset   : {dataset_display}",
        f"  Generated : {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}",
        SEP, "",
    ]

    crit      = layer_results[0]
    auroc_str = f"{crit['auroc_cont']:.3f}" if crit["auroc_cont"] is not None else "N/A"
    lines += [
        f"  Critical layer   : {crit['layer_name']}  (AUROC_cont = {auroc_str})",
        f"  Baseline pass@1  : {crit.get('baseline_p1', 0)*100:.1f}%  "
        f"(n={crit.get('n_test','?')} problems,  wrong runs={crit.get('n_wrong_total','?')},  right runs={crit.get('n_right_total','?')})",
        f"  Layers evaluated : {len(layer_results)}",
        "",
    ]

    hdr = f"  {'α':>7}  {'pass@1':>8}  {'Δ p@1':>8}"

    for res in layer_results:
        ac_s = f"{res['auroc_cont']:.3f}" if res["auroc_cont"] is not None else "N/A"
        lines += [SEP2, f"  Layer: {res['layer_name']}  (AUROC_cont = {ac_s})", SEP2]

        lines.append(
            f"  ● Experiment 1 — Wrong → Right  (α > 0)"
            f"  [{res.get('n_wrong_total','?')} wrong runs,  baseline={res.get('baseline_p1',0)*100:.1f}%]"
        )
        lines.append(hdr)
        for row in res["exp1"]:
            lines.append(
                f"  {row['alpha']:>7.1f}"
                f"  {row['pass@1']*100:>7.1f}%  {row['delta_p1']:>+7.1f}%"
            )
        lines += [f"  Best Fwd Δp@1 = {res['best_forward_p1']:+.1f}%", ""]

        lines.append(
            f"  ● Experiment 2 — Right → Wrong  (α < 0)"
            f"  [{res.get('n_right_total','?')} right runs,  baseline={res.get('baseline_p1',0)*100:.1f}%]"
        )
        lines.append(hdr)
        for row in res["exp2"]:
            lines.append(
                f"  {row['alpha']:>7.1f}"
                f"  {row['pass@1']*100:>7.1f}%  {row['delta_p1']:>+7.1f}%"
            )
        lines += [f"  Best Bwd Δp@1 = {res['best_backward_p1']:+.1f}%", ""]

    if len(layer_results) > 1:
        lines += [SEP2, "  TOP-N COMPARISON  (sorted by best Fwd Δp@1)", SEP2]
        lines.append(
            f"  {'Layer':12}  {'AUROC_cont':>10}"
            f"  {'Best Fwd Δp@1':>14}  {'Best Bwd Δp@1':>14}"
        )
        for res in sorted(layer_results, key=lambda r: -r["best_forward_p1"]):
            ac_s = f"{res['auroc_cont']:.3f}" if res["auroc_cont"] is not None else "       N/A"
            lines.append(
                f"  {res['layer_name']:12}  {ac_s:>10}"
                f"  {res['best_forward_p1']:>+13.1f}%"
                f"  {res['best_backward_p1']:>+13.1f}%"
            )
        lines.append("")

    lines.append(SEP)
    text = "\n".join(lines)
    print(text)
    path = steer_dir / "steering_report.txt"
    path.write_text(text, encoding="utf-8")
    print(f"  Report  → {path}")



def write_steering_report_xlsx(
    steer_dir:       Path,
    model_display:   str,
    dataset_display: str,
    layer_results:   list[dict],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WARN] openpyxl not installed — Excel report skipped.")
        return

    wb     = openpyxl.Workbook()
    HEAD_F = PatternFill("solid", fgColor="FFD9E1F2")
    GOOD_F = PatternFill("solid", fgColor="FFC6EFCE")
    BAD_F  = PatternFill("solid", fgColor="FFFFC7CE")
    BOLD   = Font(bold=True)
    CTR    = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "Summary"
    ws.append([f"PIPELINE-2 STEERING — {model_display} — {dataset_display}"])
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.append([])

    crit = layer_results[0]
    for label, val in [
        ("Critical layer",  crit["layer_name"]),
        ("Baseline pass@1", f"{crit.get('baseline_p1', 0)*100:.1f}%  (from stored runs)"),
        ("Test problems",   f"{crit.get('n_test','?')}  (wrong runs={crit.get('n_wrong_total','?')},  right runs={crit.get('n_right_total','?')})"),
        ("Generated",       datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]:
        ws.cell(ws.max_row + 1, 1, label).font = BOLD
        ws.cell(ws.max_row, 2, str(val))
    ws.append([])

    hdr = ["Layer", "AUROC_cont", "Baseline p@1 (%)",
           "n Wrong Runs", "Best α (fwd)", "Best p@1 fwd (%)", "Best Δp@1 fwd (%)",
           "n Right Runs", "Best α (bwd)", "Best p@1 bwd (%)", "Best Δp@1 bwd (%)"]
    ws.append(hdr)
    r_hdr = ws.max_row
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(r_hdr, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
    ws.freeze_panes = f"A{r_hdr + 1}"

    for res in layer_results:
        e1 = max(res["exp1"], key=lambda r: r["delta_p1"]) if res["exp1"] else {}
        e2 = min(res["exp2"], key=lambda r: r["delta_p1"]) if res["exp2"] else {}
        row_data = [
            res["layer_name"], res["auroc_cont"],
            round(res.get("baseline_p1", 0) * 100, 2),
            res.get("n_wrong_total", 0),
            e1.get("alpha"), round(e1.get("pass@1", 0) * 100, 2), round(e1.get("delta_p1", 0), 2),
            res.get("n_right_total", 0),
            e2.get("alpha"), round(e2.get("pass@1", 0) * 100, 2), round(e2.get("delta_p1", 0), 2),
        ]
        ws.append(row_data)
        ri = ws.max_row
        ws.cell(ri,  7).fill = GOOD_F if (row_data[6] or 0) > 0 else BAD_F
        ws.cell(ri, 11).fill = GOOD_F if (row_data[10] or 0) < 0 else BAD_F
        for ci in range(1, len(hdr) + 1):
            ws.cell(ri, ci).alignment = CTR

    for ci, w in enumerate([14, 12, 14, 12, 12, 14, 16, 12, 12, 14, 16], 1):
        ws.column_dimensions[chr(64 + ci)].width = w

    ws2 = wb.create_sheet("Exp1 Wrong→Right")
    ws2.append(["Experiment 1: Wrong → Right (α > 0)  — all contrastive test problems"])
    ws2.cell(1, 1).font = Font(bold=True, size=12)
    ws2.append([])
    hdr2 = ["Layer", "α", "pass@1 (%)", "Δp@1 (%)"]
    ws2.append(hdr2)
    for ci, h in enumerate(hdr2, 1):
        c = ws2.cell(3, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
    ws2.freeze_panes = "A4"
    for res in layer_results:
        for row in res["exp1"]:
            ws2.append([
                res["layer_name"], row["alpha"],
                round(row["pass@1"] * 100, 2), round(row["delta_p1"], 2),
            ])
            ri = ws2.max_row
            ws2.cell(ri, 4).fill = GOOD_F if row["delta_p1"] > 0 else BAD_F
            for ci in range(1, 5):
                ws2.cell(ri, ci).alignment = CTR

    ws3 = wb.create_sheet("Exp2 Right→Wrong")
    ws3.append(["Experiment 2: Right → Wrong (α < 0)  — all contrastive test problems"])
    ws3.cell(1, 1).font = Font(bold=True, size=12)
    ws3.append([])
    hdr3 = ["Layer", "α", "pass@1 (%)", "Δp@1 (%)"]
    ws3.append(hdr3)
    for ci, h in enumerate(hdr3, 1):
        c = ws3.cell(3, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
    ws3.freeze_panes = "A4"
    for res in layer_results:
        for row in res["exp2"]:
            ws3.append([
                res["layer_name"], row["alpha"],
                round(row["pass@1"] * 100, 2), round(row["delta_p1"], 2),
            ])
            ri = ws3.max_row
            ws3.cell(ri, 4).fill = GOOD_F if row["delta_p1"] < 0 else BAD_F
            for ci in range(1, 5):
                ws3.cell(ri, ci).alignment = CTR

    path = steer_dir / "steering_report.xlsx"
    wb.save(path)
    print(f"  Excel   → {path}")



def save_plots_p2(
    steer_dir:     Path,
    model_display: str,
    layer_results: list[dict],
) -> None:
    plot_dir = steer_dir / "plots"
    plot_dir.mkdir(exist_ok=True)
    crit = layer_results[0]
    bp1  = crit.get("baseline_p1", 0) * 100

    fig, axes = plt.subplots(1, 2, figsize=(12, 4))
    for ax_, exp_rows, title in [
        (axes[0], crit["exp1"], "Exp1: Wrong → Right"),
        (axes[1], crit["exp2"], "Exp2: Right → Wrong"),
    ]:
        alphas = [r["alpha"]      for r in exp_rows]
        ax_.axhline(bp1, color="steelblue", ls="--", lw=1, label=f"Baseline p@1 ({bp1:.1f}%)")
        ax_.plot(alphas, [r["pass@1"]*100 for r in exp_rows], "o-", color="darkorange", lw=1.8, label="pass@1 steered")
        ax_.set_xlabel("α"); ax_.set_ylabel("pass@1 (%)"); ax_.set_title(f"{title}\n{crit['layer_name']}")
        ax_.legend(fontsize=8); ax_.grid(alpha=0.3)
    fig.suptitle(f"Critical Layer Steering — {model_display}", fontsize=11)
    fig.tight_layout()
    fig.savefig(plot_dir / "critical_layer_steering.png", dpi=150)
    plt.close(fig)

    if len(layer_results) > 1:
        layers = [r["layer_name"]      for r in layer_results]
        d_fwd  = [r["best_forward_p1"] for r in layer_results]
        d_bwd  = [r["best_backward_p1"] for r in layer_results]
        x = np.arange(len(layers)); w = 0.35
        fig, ax = plt.subplots(figsize=(max(6, len(layers) * 1.2), 4))
        ax.bar(x - w/2, d_fwd, w, color="steelblue", label="Best Fwd Δp@1")
        ax.bar(x + w/2, d_bwd, w, color="tomato",    label="Best Bwd Δp@1")
        ax.axhline(0, color="black", lw=0.7)
        ax.set_xticks(x); ax.set_xticklabels(layers, rotation=30, ha="right", fontsize=8)
        ax.set_ylabel("Δ pass@1 (%)"); ax.set_title(f"Best Δp@1 per Layer — {model_display}")
        ax.legend(fontsize=9); ax.grid(axis="y", alpha=0.3); fig.tight_layout()
        fig.savefig(plot_dir / "layer_delta_p1.png", dpi=150)
        plt.close(fig)

        colors = plt.cm.tab10(np.linspace(0, 1, len(layer_results)))
        fig, axes = plt.subplots(1, 2, figsize=(13, 4))
        for res, c in zip(layer_results, colors):
            axes[0].plot([r["alpha"] for r in res["exp1"]],
                         [r["pass@1"]*100 for r in res["exp1"]], "o-", color=c, label=res["layer_name"])
            axes[1].plot([r["alpha"] for r in res["exp2"]],
                         [r["pass@1"]*100 for r in res["exp2"]], "s--", color=c, label=res["layer_name"])
        for ax_ in axes:
            ax_.axhline(bp1, color="gray", ls="--", lw=1, label=f"Baseline ({bp1:.1f}%)")
            ax_.set_xlabel("α"); ax_.set_ylabel("pass@1 (%)"); ax_.grid(alpha=0.3); ax_.legend(fontsize=7)
        axes[0].set_title("Exp1: Wrong → Right")
        axes[1].set_title("Exp2: Right → Wrong")
        fig.suptitle(f"All Layers pass@1 — {model_display}", fontsize=11)
        fig.tight_layout()
        fig.savefig(plot_dir / "all_layers_p1.png", dpi=150)
        plt.close(fig)

    print(f"  Plots   → {plot_dir}")



def generate_dataset_combined_report(
    ds_results_dir: Path,
    dataset_key:    str,
    all_model_data: list[dict],
) -> None:
    """
    Generate <dataset>_report.txt and <dataset>_report.xlsx combining
    probing + steering summaries for all models.

    Each entry of all_model_data:
      {"model_key", "model_display",
       "probe_analysis", "top_layers",   # from probing
       "steering"}                        # from steering
    """
    if not all_model_data:
        return

    ds_display = _DS_DISPLAY.get(dataset_key, dataset_key)
    SEP  = "=" * 84
    SEP2 = "-" * 84
    lines: list[str] = []

    lines += [
        SEP,
        f"  PIPELINE-2 COMBINED REPORT — {ds_display}",
        f"  Models   : {', '.join(m['model_display'] for m in all_model_data)}",
        f"  Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}",
        SEP, "",
    ]

    lines += ["  PROBING SUMMARY", SEP2]
    lines.append(
        f"  {'Model':25}  {'Best AUROC_cont':>15}  {'Top Layer':>10}"
    )
    lines.append("  " + "─" * 56)
    for m in all_model_data:
        pa   = m.get("probe_analysis", [])
        tops = m.get("top_layers", [])
        bac  = max((r["auroc_cont"] for r in pa if r.get("auroc_cont") is not None), default=None)
        tl   = tops[0]["layer_name"] if tops else "N/A"
        bac_s = f"{bac:.3f}" if bac is not None else "        N/A"
        lines.append(f"  {m['model_display']:25}  {bac_s:>15}  {tl:>10}")
    lines.append("")

    lines += ["  STEERING SUMMARY  (metric: pass@1)", SEP2]
    lines.append(
        f"  {'Model':25}  {'Baseline p@1':>13}  {'Wrong Runs':>10}  {'Right Runs':>10}"
        f"  {'Best Fwd Δp@1':>14}  {'Best Bwd Δp@1':>14}"
    )
    lines.append("  " + "─" * 92)
    for m in all_model_data:
        steer = m.get("steering", [])
        if not steer:
            lines.append(f"  {m['model_display']:25}  {'N/A':>13}  {'':>10}  {'':>10}"
                         f"  {'N/A':>14}  {'N/A':>14}")
            continue
        bp1  = steer[0].get("baseline_p1", 0) * 100
        nw   = steer[0].get("n_wrong_total", 0)
        nr   = steer[0].get("n_right_total", 0)
        bfp1 = max((r.get("best_forward_p1",  0) for r in steer), default=0)
        bbp1 = min((r.get("best_backward_p1", 0) for r in steer), default=0)
        lines.append(
            f"  {m['model_display']:25}  {bp1:>12.1f}%  {nw:>10}  {nr:>10}"
            f"  {bfp1:>+13.1f}%  {bbp1:>+13.1f}%"
        )
    lines += ["", SEP]

    text = "\n".join(lines)
    print(text)
    txt_path = ds_results_dir / f"{dataset_key}_report.txt"
    txt_path.write_text(text, encoding="utf-8")
    print(f"  Dataset report    → {txt_path}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return

    wb     = openpyxl.Workbook()
    HEAD_F = PatternFill("solid", fgColor="FFD9E1F2")
    GOOD_F = PatternFill("solid", fgColor="FFC6EFCE")
    BAD_F  = PatternFill("solid", fgColor="FFFFC7CE")
    BEST_F = PatternFill("solid", fgColor="FFFFD700")
    BOLD   = Font(bold=True)
    CTR    = Alignment(horizontal="center")

    def _write_sheet(ws, title, hdr, rows_data, color_cols=None):
        ws.append([title]); ws.cell(1, 1).font = Font(bold=True, size=13); ws.append([])
        ws.append(hdr)
        for ci, h in enumerate(hdr, 1):
            c = ws.cell(3, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
        ws.freeze_panes = "A4"
        for row in rows_data:
            ws.append(row)
            ri = ws.max_row
            for ci in range(1, len(hdr) + 1):
                ws.cell(ri, ci).alignment = CTR
            if color_cols:
                for ci, val in color_cols(row):
                    ws.cell(ri, ci).fill = GOOD_F if val > 0 else BAD_F

    ws = wb.active
    ws.title = "Probing"
    p_rows = []
    for m in all_model_data:
        pa   = m.get("probe_analysis", [])
        tops = m.get("top_layers", [])
        baa  = max((r["auroc_all"]  for r in pa if r.get("auroc_all")  is not None), default=None)
        bac  = max((r["auroc_cont"] for r in pa if r.get("auroc_cont") is not None), default=None)
        tl   = tops[0]["layer_name"] if tops else "N/A"
        p_rows.append([m["model_display"], baa, bac, tl, len(tops)])
    _write_sheet(ws, f"PROBING SUMMARY — {ds_display}",
                 ["Model", "Best AUROC_all", "Best AUROC_cont", "Top Layer (1st)", "# Layers"],
                 p_rows)
    ws.column_dimensions["A"].width = 28
    for col in "BCDE":
        ws.column_dimensions[col].width = 16

    ws2 = wb.create_sheet("Steering")
    s_rows = []
    for m in all_model_data:
        steer = m.get("steering", [])
        if not steer:
            s_rows.append([m["model_display"]] + [None] * 5)
            continue
        bp1  = round(steer[0].get("baseline_p1", 0) * 100, 2)
        nw   = steer[0].get("n_wrong_total", 0)
        nr   = steer[0].get("n_right_total", 0)
        bfp1 = round(max((r.get("best_forward_p1",  0) for r in steer), default=0), 2)
        bbp1 = round(min((r.get("best_backward_p1", 0) for r in steer), default=0), 2)
        s_rows.append([m["model_display"], bp1, nw, nr, bfp1, bbp1])

    def _steer_colors(row):
        if len(row) > 4 and row[4] is not None: yield 5, row[4]
        if len(row) > 5 and row[5] is not None: yield 6, -abs(row[5])

    _write_sheet(ws2, f"STEERING SUMMARY — {ds_display}",
                 ["Model", "Baseline p@1 (%)", "n Wrong Runs", "n Right Runs",
                  "Best Fwd Δp@1 (%)", "Best Bwd Δp@1 (%)"],
                 s_rows, color_cols=_steer_colors)
    ws2.column_dimensions["A"].width = 28
    for col in "BCDEFG":
        ws2.column_dimensions[col].width = 18

    xlsx_path = ds_results_dir / f"{dataset_key}_report.xlsx"
    wb.save(xlsx_path)
    print(f"  Dataset Excel     → {xlsx_path}")



def run_one_model_p2(
    model_key:   str,
    acts_dir:    str,
    results_dir: Path,
    ds_key:      str,
    alphas_fwd:  list[float],
    alphas_bwd:  list[float],
    n_gen:       int,
    device:      str,
    apps_difficulty: str = "introductory",
    apps_max:    int = 500,
    problem_map: dict | None = None,
) -> list[dict] | None:
    """
    Run pipeline-2 steering for one model.

    Parameters
    ----------
    acts_dir    : activation run directory (split.json + H5 files)
    results_dir : pipeline2 results directory for this model run
                  (must contain probing/top_layers.json + probing/directions/)

    Returns
    -------
    list of layer result dicts (stripped of per_problem), or None on failure
    """
    model_cfg  = MODEL_REGISTRY[model_key]
    steer_dir  = results_dir / "steering"
    done_flag  = steer_dir / "pipeline2_results.json"

    if done_flag.exists():
        print(f"  [SKIP] {model_cfg['display']} — steering already done.")
        with open(done_flag) as f:
            return json.load(f)

    steer_dir.mkdir(parents=True, exist_ok=True)
    ds_display = _DS_DISPLAY.get(ds_key, ds_key)

    print("\n" + "=" * 76)
    print(f"  PIPELINE-2 STEERING — {model_cfg['display']}  [{ds_display}]")
    print(f"  acts_dir    : {acts_dir}")
    print(f"  results_dir : {results_dir}")
    print(f"  αs fwd  : {alphas_fwd}")
    print(f"  αs bwd  : {[-a for a in alphas_bwd]}")
    print("=" * 76)

    split_path = Path(acts_dir) / "split.json"
    if not split_path.exists():
        print(f"  [ERROR] split.json not found in {acts_dir}")
        return None

    with open(split_path) as f:
        split = json.load(f)

    test_ids = set(split["test"]["contrastive_ids"])
    print(f"  Contrastive test : {len(test_ids)} problems")

    if not test_ids:
        print("  [SKIP] No contrastive test IDs.")
        return None

    try:
        top_layers = load_top_layers(results_dir)
        directions = load_directions(results_dir)
    except FileNotFoundError as exc:
        print(f"  [ERROR] {exc}")
        return None

    layers_to_run = [rec for rec in top_layers if rec["layer_name"] in directions]
    skipped       = [rec["layer_name"] for rec in top_layers if rec["layer_name"] not in directions]
    if skipped:
        print(f"  [WARN] Direction files missing for: {skipped}")
    if not layers_to_run:
        print("  [ERROR] No directions available.")
        return None

    print(
        "  Top layers: "
        + ", ".join(
            r["layer_name"]
            + (f"({r['auroc_cont']:.3f})" if r.get("auroc_cont") is not None else "")
            for r in layers_to_run
        )
    )

    if problem_map is None:
        problem_map = load_problems_with_apps_fix(
            ds_key, test_ids,
            apps_difficulty=apps_difficulty,
            apps_max=apps_max,
        )

    test_problems = {tid: problem_map[tid] for tid in test_ids if tid in problem_map}
    missing = test_ids - set(test_problems.keys())
    if missing:
        print(f"  WARNING: {len(missing)} test IDs not in dataset — skipped.")
        for tid in sorted(missing)[:10]:
            print(f"    {tid}")
        if len(missing) > 10:
            print(f"    … and {len(missing) - 10} more")
    print(f"  Test problems: {len(test_problems)}/{len(test_ids)}")

    if not test_problems:
        print("  [SKIP] No test problems.")
        return None

    run_counts = compute_baseline(acts_dir, test_ids)
    n_wrong_total = sum(v["n_wrong"] for v in run_counts.values())
    n_right_total = sum(v["n_right"] for v in run_counts.values())
    print(f"  Contrastive test: {len(run_counts)} problems  "
          f"(wrong runs={n_wrong_total}  right runs={n_right_total})")
    print(f"  Exp1 will steer {n_wrong_total} wrong runs  →  want improvement above baseline")
    print(f"  Exp2 will steer {n_right_total} right runs  →  want degradation below baseline")

    model, tokenizer = load_model(model_cfg, device)

    all_layer_results: list[dict] = []

    try:
        n_model_layers = len(_get_decoder_layer(model, 0).__class__.__mro__)
        import torch.nn as _nn
        for _attr in ("model", "transformer", "gpt_neox", "language_model"):
            _sub = getattr(model, _attr, None)
            if _sub is not None:
                for _lattr in ("layers", "h", "blocks"):
                    _layers_obj = getattr(_sub, _lattr, None)
                    if isinstance(_layers_obj, _nn.ModuleList):
                        n_model_layers = len(_layers_obj)
                        break
                else:
                    continue
                break
    except Exception:
        n_model_layers = None

    valid_layers = []
    for layer_info in layers_to_run:
        li = layer_info["layer_idx"]
        if n_model_layers is not None and li >= n_model_layers:
            print(f"  [SKIP] {layer_info['layer_name']} (idx={li}) — out of range "
                  f"for model with {n_model_layers} layers.")
        else:
            valid_layers.append(layer_info)

    if len(valid_layers) < len(layers_to_run):
        print(f"  [INFO] {len(layers_to_run) - len(valid_layers)} layer(s) skipped, "
              f"{len(valid_layers)} remaining.")

    for rank, layer_info in enumerate(
        tqdm(valid_layers, desc=f"Layers [{model_cfg['display']}]", unit="layer"), 1
    ):
        lname        = layer_info["layer_name"]
        li           = layer_info["layer_idx"]
        auroc_cont   = layer_info.get("auroc_cont")
        direction_np = directions[lname]

        tqdm.write(
            f"\n  ── Layer {rank}/{len(valid_layers)}: {lname}  "
            f"(AUROC_cont={'N/A' if auroc_cont is None else f'{auroc_cont:.3f}'}) ──"
        )

        try:
            layer_result = run_layer_experiment_p2(
                model, tokenizer, model_cfg,
                test_problems, direction_np,
                li, lname, auroc_cont,
                run_counts,
                alphas_fwd, alphas_bwd,
                device,
            )
        except Exception as _layer_exc:
            import traceback as _tb
            tqdm.write(f"  [ERROR] {lname} failed: {_layer_exc} — skipping layer.")
            _tb.print_exc()
            continue

        all_layer_results.append(layer_result)

        checkpoint = {
            k: (_strip_per_problem(v) if k in ("exp1", "exp2") else v)
            for k, v in layer_result.items()
        }
        with open(steer_dir / f"{lname}_results.json", "w") as f:
            json.dump(checkpoint, f, indent=2)

    unload_model(model, tokenizer)
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

    full_summary = [
        {k: (_strip_per_problem(v) if k in ("exp1", "exp2") else v)
         for k, v in res.items()}
        for res in all_layer_results
    ]
    with open(done_flag, "w") as f:
        json.dump(full_summary, f, indent=2)

    with open(steer_dir / "steering_summary.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "layer", "auroc_cont", "baseline_p1", "n_test",
            "best_alpha_fwd", "best_p1_fwd", "best_fwd_delta_p1",
            "best_alpha_bwd", "best_p1_bwd", "best_bwd_delta_p1",
        ])
        for res in all_layer_results:
            e1 = max(res["exp1"], key=lambda r: r["delta_p1"]) if res["exp1"] else {}
            e2 = min(res["exp2"], key=lambda r: r["delta_p1"]) if res["exp2"] else {}
            writer.writerow([
                res["layer_name"],
                f"{res['auroc_cont']:.4f}" if res["auroc_cont"] is not None else "",
                f"{res.get('baseline_p1', 0):.4f}", res.get("n_test", 0),
                f"{e1.get('alpha', '')}", f"{e1.get('pass@1', 0):.4f}" if e1 else "",
                f"{e1.get('delta_p1', 0):.2f}" if e1 else "",
                f"{e2.get('alpha', '')}", f"{e2.get('pass@1', 0):.4f}" if e2 else "",
                f"{e2.get('delta_p1', 0):.2f}" if e2 else "",
            ])

    write_steering_report_txt(steer_dir, model_cfg["display"], ds_display, all_layer_results)
    write_steering_report_xlsx(steer_dir, model_cfg["display"], ds_display, all_layer_results)
    try:
        save_plots_p2(steer_dir, model_cfg["display"], all_layer_results)
    except Exception as exc:
        print(f"  [WARN] Plotting failed: {exc}")

    print(f"\n  Saved → {steer_dir}")
    return full_summary



def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Pipeline-2 steering: steer model using pre-computed directions from probing.py.\n"
            "All outputs go to PIPELINE_2_RESULTS/<dataset>/<model>_<ts>/steering/."
        )
    )
    parser.add_argument("--dataset",      default="mbppplus", choices=["mbppplus"])
    parser.add_argument("--model",        default=None, choices=list(MODEL_REGISTRY.keys()))
    parser.add_argument("--acts-dir",     default=None,
                        help="Activation run directory (split.json + H5 files).")
    parser.add_argument("--results-dir",  default=None,
                        help="Probing results dir for this model (from probing.py).")
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT,
                        help="Root for results (default: ../mbppplus).")
    parser.add_argument("--alphas-fwd",   nargs="+", type=float, default=DEFAULT_ALPHAS_FWD)
    parser.add_argument("--alphas-bwd",   nargs="+", type=float, default=DEFAULT_ALPHAS_BWD,
                        help="α magnitudes for Exp2 (negated internally).")
    parser.add_argument("--n-gen",        type=int, default=N_STEER_RUNS)
    parser.add_argument("--device",       default=None)
    args = parser.parse_args()

    device       = args.device or ("cuda" if torch.cuda.is_available() else "cpu")
    alphas_bwd   = [-abs(a) for a in args.alphas_bwd]
    results_root = Path(args.results_root).expanduser()

    kw = dict(
        alphas_fwd=args.alphas_fwd, alphas_bwd=alphas_bwd,
        n_gen=args.n_gen, device=device,
        apps_difficulty="introductory", apps_max=500,
    )

    if args.acts_dir and args.model and args.dataset and args.results_dir:
        results_dir = Path(args.results_dir)
        if not (results_dir / "probing" / "top_layers.json").exists():
            sys.exit(f"[ERROR] Probing not done in {results_dir}. Run probing.py first.")
        run_one_model_p2(
            model_key=args.model, acts_dir=args.acts_dir,
            results_dir=results_dir, ds_key=args.dataset, **kw,
        )
        return

    if not results_root.exists():
        print(f"[ERROR] PIPELINE_2_RESULTS root not found: {results_root}")
        print("Run probing.py first.")
        return

    ds_key = "mbppplus"
    runs   = _find_p2_runs_with_probing(results_root, ds_key)

    print("\n" + "=" * 66)
    print("  CODE_LLM STEERING — MBPP+ PROBING-COMPLETE RUNS")
    print("=" * 66)
    if runs:
        for i, (mk, rd) in enumerate(runs, 1):
            label = MODEL_REGISTRY[mk]["display"] if mk in MODEL_REGISTRY else mk
            print(f"  {i}. {label:<28}  ←  {rd.name}")
    else:
        print("  No probing-complete runs found.")
        print("  Run: python probing.py")
    print("=" * 66)

    if not runs:
        return

    available = [(ds_key, runs)]

    ds_results_dir = results_root / ds_key

    print(f"\n  {len(runs)} model(s) with completed probing in {_DS_DISPLAY.get(ds_key, ds_key)}:")
    for i, (mk, rd) in enumerate(runs, 1):
        label = MODEL_REGISTRY[mk]["display"] if mk in MODEL_REGISTRY else mk
        print(f"    {i}. {label}  ←  {rd.name}")

    print(f"\n  Select model(s) to steer [1-{len(runs)}, comma-separated, or Enter for all]: ", end="", flush=True)
    raw = input().strip()
    if raw:
        selected_indices = []
        for part in raw.split(","):
            part = part.strip()
            if part.isdigit() and 1 <= int(part) <= len(runs):
                selected_indices.append(int(part) - 1)
            else:
                print(f"  [WARN] Ignoring invalid selection: {part!r}")
        if selected_indices:
            runs = [runs[i] for i in selected_indices]
        else:
            print("  No valid selection — running all models.")
    print(f"  Running steering for: {', '.join(MODEL_REGISTRY[mk]['display'] if mk in MODEL_REGISTRY else mk for mk, _ in runs)}")

    print("\n  Collecting all contrastive test IDs …")
    all_test_ids: set[str] = set()
    acts_dir_map: dict[str, str] = {}

    for mk, results_dir in runs:
        try:
            pa = load_probe_analysis(results_dir)
            acts_dir = pa.get("acts_dir", "")
            acts_dir_map[mk] = acts_dir
            sp = Path(acts_dir) / "split.json"
            if sp.exists():
                with open(sp) as f:
                    sp_data = json.load(f)
                all_test_ids.update(sp_data.get("test", {}).get("contrastive_ids", []))
        except Exception as exc:
            print(f"  [WARN] Could not read acts_dir for {mk}: {exc}")

    try:
        problem_map: dict | None = load_problems_with_apps_fix(
            ds_key, all_test_ids,
            apps_difficulty=args.apps_difficulty,
            apps_max=args.apps_max,
        )
    except Exception as exc:
        print(f"  [WARN] Could not pre-load dataset ({exc}). Will load per-model.")
        problem_map = None

    all_model_data: list[dict] = []

    for i, (mk, results_dir) in enumerate(runs, 1):
        label = MODEL_REGISTRY[mk]["display"] if mk in MODEL_REGISTRY else mk
        print(f"\n{'#'*76}")
        print(f"  Model {i}/{len(runs)}: {label}")
        print(f"{'#'*76}")

        acts_dir = acts_dir_map.get(mk, "")
        if not acts_dir or not Path(acts_dir).exists():
            print(f"  [WARN] acts_dir not found ({acts_dir}) — trying probe_analysis.json …")
            try:
                pa       = load_probe_analysis(results_dir)
                acts_dir = pa.get("acts_dir", "")
            except Exception:
                pass
        if not acts_dir or not Path(acts_dir).exists():
            print(f"  [ERROR] Cannot determine activation dir for {mk} — skipping.")
            continue

        try:
            pa       = load_probe_analysis(results_dir)
            top_lays = load_top_layers(results_dir)
        except Exception as exc:
            print(f"  [WARN] Could not read probing data: {exc}")
            pa = {}; top_lays = []

        try:
            steer_results = run_one_model_p2(
                model_key=mk, acts_dir=acts_dir,
                results_dir=results_dir, ds_key=ds_key,
                problem_map=problem_map, **kw,
            )
        except Exception as exc:
            import traceback
            print(f"\n  [ERROR] {label} failed: {exc}")
            traceback.print_exc()
            steer_results = None

        if not steer_results:
            steer_dir_fallback = results_dir / "steering"
            recovered = []
            if steer_dir_fallback.exists():
                for ckpt in sorted(steer_dir_fallback.glob("layer_*_results.json")):
                    try:
                        with open(ckpt) as _f:
                            recovered.append(json.load(_f))
                    except Exception:
                        pass
            if recovered:
                print(f"  [INFO] Recovered {len(recovered)} layer result(s) from checkpoints.")
                steer_results = recovered

        all_model_data.append({
            "model_key":      mk,
            "model_display":  label,
            "probe_analysis": pa.get("all_layers", []),
            "top_layers":     top_lays,
            "steering":       steer_results or [],
        })

    if all_model_data:
        print(f"\n{'='*76}")
        print("  Generating dataset combined report …")
        generate_dataset_combined_report(ds_results_dir, ds_key, all_model_data)


if __name__ == "__main__":
    main()
