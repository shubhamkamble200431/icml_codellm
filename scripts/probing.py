"""
CODE_LLM probing.py  —  Step 2: Probe analysis on stored activations
=========================================================================
Reads layer-wise H5 activations produced by contrastive_set_gen.py and
runs a probe analysis to select the best steering layers.

For every layer:
  1. Probe-all   — trained on all stored activations (excl. cont-test)
  2. Probe-cont  — trained on contrastive-train only
  3. gap         = AUROC_all − AUROC_cont
  4. If gap > threshold → use contrastive-only direction
  5. Bootstrapped direction (10 × downsampled, averaged, normalised)
  6. Top-N layer selection: sorted by AUROC_cont descending

Output structure
----------------
All outputs go to  ../mbppplus/<model_key>_<ts>/probing/

  probe_analysis.json    — per-layer stats + path to acts_dir
  top_layers.json        — selected top-N metadata
  directions/layer_XX.npy
  plots/
  probing_report.txt
  probing_report.xlsx

Usage
-----
  # Interactive (auto-discovers MBPP+ activation runs)
  python probing.py

  # CLI single model
  python probing.py \\
      --model    qwen-coder-1.5b-instruct \\
      --acts-dir /path/to/mbppplus/qwen-coder-1.5b-instruct_xxx \\
      [--top-n 5] [--n-bootstrap 10] [--gap-thresh 0.15] [--rng-seed 42]
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import h5py
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import roc_auc_score
from tqdm import tqdm

import importlib.util as _ilu

_GEN_PATH = Path(__file__).parent / "contrastive_set_gen.py"
_spec     = _ilu.spec_from_file_location("_gen", _GEN_PATH)
_gen      = _ilu.module_from_spec(_spec)
_spec.loader.exec_module(_gen)

MODEL_REGISTRY = _gen.MODEL_REGISTRY

_DS_DISPLAY = {
    "mbppplus": "MBPP+  (~378 problems, more robust tests — evalplus)",
}


def _infer_model_key(dir_name: str) -> str | None:
    for mk in MODEL_REGISTRY:
        if dir_name.startswith(mk + "_") or dir_name == mk:
            return mk
    return None


_MODEL_KEY_ORDER = [
    "qwen-coder-1.5b-instruct",
    "qwen-coder-7b-instruct",
]



N_BOOTSTRAP  = 10
TOP_N_LAYERS = 5
GAP_THRESH   = 0.15
RNG_SEED     = 42

SKIP_DIRS = {
    "probing", "plots", "steering",
    "pipeline2_steering", "pipeline2_probing", "pipeline2_results",
    "all", "contrastive", "comparison",
}

DEFAULT_BASE         = str(Path(__file__).parent.parent / "mbppplus")
DEFAULT_RESULTS_ROOT = str(Path(__file__).parent.parent / "mbppplus")



_APPS_REQUIRED_DIFFICULTY = "competition"


def _read_apps_difficulty(run_dir: Path) -> str | None:
    """Return apps_difficulty from split.json, or None if not present."""
    sp = run_dir / "split.json"
    if not sp.exists():
        return None
    try:
        with open(sp) as f:
            return json.load(f).get("apps_difficulty")
    except Exception:
        return None


def _find_activation_runs(ds_dir: Path) -> list[tuple[str, Path]]:
    """Return [(model_key, acts_dir), ...] for completed activation run dirs.

    For APPS, only runs with apps_difficulty == 'competition' are returned.
    """
    best: dict[str, Path] = {}
    for d in sorted(ds_dir.iterdir()):
        if not d.is_dir() or not (d / "split.json").exists():
            continue
        mk = _infer_model_key(d.name)
        if mk is None:
            continue
        if ds_dir.name == "apps":
            diff = _read_apps_difficulty(d)
            if diff != _APPS_REQUIRED_DIFFICULTY:
                continue
        best[mk] = d
    ordered  = [(mk, best[mk]) for mk in _MODEL_KEY_ORDER if mk in best]
    ordered += [(mk, best[mk]) for mk in best if mk not in _MODEL_KEY_ORDER]
    return ordered


def _find_p2_run(results_root: Path, dataset: str, model_key: str,
                 require_file: str | None = None) -> Path | None:
    """Return most-recent pipeline2 results dir for this model, or None."""
    ds_dir = results_root / dataset
    if not ds_dir.exists():
        return None
    candidates = [
        d for d in sorted(ds_dir.iterdir())
        if d.is_dir() and d.name.startswith(model_key + "_")
    ]
    if require_file:
        candidates = [d for d in candidates if (d / require_file).exists()]
    return candidates[-1] if candidates else None


def _new_results_dir(results_root: Path, dataset: str, model_key: str) -> Path:
    """Create and return a fresh timestamped results dir."""
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    d   = results_root / dataset / f"{model_key}_{ts}"
    d.mkdir(parents=True, exist_ok=True)
    return d



def _safe_id(tid: str) -> str:
    return tid.replace("/", "_")


def _find_search_root(acts_dir: str) -> Path:
    root = Path(acts_dir)
    return root / "all" if (root / "all").is_dir() else root


def _load_h5(path: Path) -> np.ndarray:
    with h5py.File(path, "r") as hf:
        return hf["activation"][:]


def detect_n_layers(acts_dir: str) -> int:
    root = _find_search_root(acts_dir)
    for prob_dir in sorted(root.iterdir()):
        if not prob_dir.is_dir() or prob_dir.name in SKIP_DIRS:
            continue
        for verdict in ("right", "wrong"):
            vdir = prob_dir / verdict
            if not vdir.exists():
                continue
            for run_dir in sorted(vdir.iterdir()):
                if run_dir.is_dir():
                    h5s = sorted(run_dir.glob("layer_*.h5"))
                    if h5s:
                        return len(h5s)
    raise RuntimeError(f"No layer H5 files found under {acts_dir}")



def collect_acts_for_ids(
    acts_dir: str,
    task_ids: set,
    layer_idx: int,
) -> tuple[np.ndarray | None, np.ndarray | None]:
    root     = _find_search_root(acts_dir)
    safe_set = {_safe_id(tid) for tid in task_ids}
    fname    = f"layer_{layer_idx:02d}.h5"
    rights, wrongs = [], []

    for prob_dir in sorted(root.iterdir()):
        if not prob_dir.is_dir() or prob_dir.name not in safe_set:
            continue
        for verdict, buf in (("right", rights), ("wrong", wrongs)):
            vdir = prob_dir / verdict
            if not vdir.exists():
                continue
            for run_dir in sorted(vdir.iterdir()):
                h5 = run_dir / fname
                if run_dir.is_dir() and h5.exists():
                    try:
                        buf.append(_load_h5(h5))
                    except Exception:
                        pass

    if not rights or not wrongs:
        return None, None
    return (np.stack(rights).astype(np.float32),
            np.stack(wrongs).astype(np.float32))


def collect_all_acts(
    acts_dir: str,
    layer_idx: int,
    exclude_ids: set | None = None,
) -> tuple[np.ndarray | None, np.ndarray | None]:
    root      = _find_search_root(acts_dir)
    excl_safe = {_safe_id(tid) for tid in (exclude_ids or set())}
    fname     = f"layer_{layer_idx:02d}.h5"
    rights, wrongs = [], []

    for prob_dir in sorted(root.iterdir()):
        if not prob_dir.is_dir() or prob_dir.name in SKIP_DIRS:
            continue
        if prob_dir.name in excl_safe:
            continue
        for verdict, buf in (("right", rights), ("wrong", wrongs)):
            vdir = prob_dir / verdict
            if not vdir.exists():
                continue
            for run_dir in sorted(vdir.iterdir()):
                h5 = run_dir / fname
                if run_dir.is_dir() and h5.exists():
                    try:
                        buf.append(_load_h5(h5))
                    except Exception:
                        pass

    if not rights or not wrongs:
        return None, None
    return (np.stack(rights).astype(np.float32),
            np.stack(wrongs).astype(np.float32))



def compute_direction_bootstrapped(
    right: np.ndarray,
    wrong: np.ndarray,
    n_samples: int = N_BOOTSTRAP,
    seed: int = RNG_SEED,
) -> tuple[np.ndarray, float, int]:
    rng = np.random.default_rng(seed)
    k   = min(len(right), len(wrong))

    if k < 2:
        diff = right.mean(0) - wrong.mean(0)
        norm = float(np.linalg.norm(diff))
        unit = (diff / norm).astype(np.float32) if norm > 1e-9 else diff.astype(np.float32)
        return unit, norm, k

    dirs = []
    for _ in range(n_samples):
        ri = rng.choice(len(right), k, replace=False)
        wi = rng.choice(len(wrong), k, replace=False)
        dirs.append(right[ri].mean(0) - wrong[wi].mean(0))

    mean_d = np.stack(dirs).mean(0)
    norm   = float(np.linalg.norm(mean_d))
    unit   = (mean_d / norm).astype(np.float32) if norm > 1e-9 else mean_d.astype(np.float32)
    return unit, norm, k



def train_probe_auroc(
    X_train: np.ndarray,
    y_train: np.ndarray,
    X_test:  np.ndarray,
    y_test:  np.ndarray,
) -> float | None:
    if len(np.unique(y_test)) < 2:
        return None

    scaler = StandardScaler()
    Xtr    = scaler.fit_transform(X_train)
    Xte    = scaler.transform(X_test)
    clf    = LogisticRegression(
        class_weight="balanced", max_iter=1000, C=1.0, solver="lbfgs"
    )
    clf.fit(Xtr, y_train)
    return float(roc_auc_score(y_test, clf.predict_proba(Xte)[:, 1]))



def select_top_layers(
    all_results: list[dict],
    top_n: int,
) -> list[dict]:
    """Deterministically pick the top-N layers by AUROC_cont (descending)."""
    valid   = [r for r in all_results if r.get("auroc_cont") is not None]
    if not valid:
        valid = all_results
    ranked  = sorted(valid, key=lambda r: r.get("auroc_cont") or 0.0, reverse=True)
    return sorted(ranked[:top_n], key=lambda r: r["layer_idx"])



def analyze_layers(
    acts_dir: str,
    split:    dict,
    top_n:       int   = TOP_N_LAYERS,
    gap_thresh:  float = GAP_THRESH,
    n_bootstrap: int   = N_BOOTSTRAP,
    seed:        int   = RNG_SEED,
) -> tuple[list[dict], list[dict]]:
    train_ids = set(split["train"]["contrastive_ids"])
    test_ids  = set(split["test"]["contrastive_ids"])
    all_cont  = train_ids | test_ids

    n_layers = detect_n_layers(acts_dir)
    tqdm.write(f"  Layers: {n_layers}  |  cont-train: {len(train_ids)}"
               f"  cont-test: {len(test_ids)}")

    all_results: list[dict] = []

    for li in tqdm(range(n_layers), desc="  Layer analysis", unit="layer"):
        lname = f"layer_{li:02d}"

        all_r, all_w = collect_all_acts(acts_dir, li, exclude_ids=test_ids)
        if all_r is None or all_w is None:
            tqdm.write(f"  [WARN] {lname}: no all-activations found — skipping")
            continue

        ct_r, ct_w = collect_acts_for_ids(acts_dir, train_ids, li)
        te_r, te_w = collect_acts_for_ids(acts_dir, test_ids,  li)

        if ct_r is not None and ct_w is not None:
            X_tr_all  = np.vstack([all_r, ct_r, all_w, ct_w])
            y_tr_all  = np.concatenate([
                np.ones(len(all_r) + len(ct_r)),
                np.zeros(len(all_w) + len(ct_w)),
            ])
            X_tr_cont = np.vstack([ct_r, ct_w])
            y_tr_cont = np.concatenate([np.ones(len(ct_r)), np.zeros(len(ct_w))])
        else:
            X_tr_all  = np.vstack([all_r, all_w])
            y_tr_all  = np.concatenate([np.ones(len(all_r)), np.zeros(len(all_w))])
            X_tr_cont = X_tr_all
            y_tr_cont = y_tr_all

        auroc_all = auroc_cont = None
        if te_r is not None and te_w is not None:
            X_te = np.vstack([te_r, te_w])
            y_te = np.concatenate([np.ones(len(te_r)), np.zeros(len(te_w))])
            try:
                auroc_all  = train_probe_auroc(X_tr_all,  y_tr_all,  X_te, y_te)
            except Exception as exc:
                tqdm.write(f"  [WARN] {lname} probe-all: {exc}")
            try:
                auroc_cont = train_probe_auroc(X_tr_cont, y_tr_cont, X_te, y_te)
            except Exception as exc:
                tqdm.write(f"  [WARN] {lname} probe-cont: {exc}")
        else:
            tqdm.write(f"  [WARN] {lname}: no cont-test activations — AUROCs=None")

        gap = None
        if auroc_all is not None and auroc_cont is not None:
            gap = round(auroc_all - auroc_cont, 4)

        use_cont_dir = (gap is not None and gap > gap_thresh)
        if use_cont_dir and ct_r is not None and ct_w is not None:
            src_r, src_w, dir_source = ct_r, ct_w, "contrastive"
        else:
            full_r, full_w = collect_acts_for_ids(acts_dir, all_cont, li)
            src_r = full_r if full_r is not None else all_r
            src_w = full_w if full_w is not None else all_w
            if not use_cont_dir:
                src_r, src_w, dir_source = all_r, all_w, "all"
            else:
                dir_source = "contrastive" if full_r is not None else "all"

        direction, raw_norm, bootstrap_k = compute_direction_bootstrapped(
            src_r, src_w, n_samples=n_bootstrap, seed=seed,
        )

        all_results.append({
            "layer_idx":    li,
            "layer_name":   lname,
            "auroc_all":    round(auroc_all,  4) if auroc_all  is not None else None,
            "auroc_cont":   round(auroc_cont, 4) if auroc_cont is not None else None,
            "gap":          gap,
            "use_cont_dir": use_cont_dir,
            "dir_source":   dir_source,
            "raw_norm":     float(raw_norm),
            "bootstrap_k":  bootstrap_k,
            "n_right_all":  int(len(all_r)),
            "n_wrong_all":  int(len(all_w)),
            "_direction":   direction,
        })

    if not all_results:
        raise RuntimeError("No layer activations could be loaded.")

    top_layers = select_top_layers(all_results, top_n)
    tqdm.write(
        f"  Top-{top_n} layers (by AUROC_cont): "
        + ", ".join(
            f"{r['layer_name']}"
            + (f"({r['auroc_cont']:.3f})" if r["auroc_cont"] is not None else "(N/A)")
            for r in top_layers
        )
    )
    return all_results, top_layers





def write_probing_report_txt(
    probe_dir:       Path,
    model_display:   str,
    dataset_display: str,
    all_results:     list[dict],
    top_layers:      list[dict],
) -> None:
    SEP  = "=" * 76
    SEP2 = "-" * 76
    lines: list[str] = []

    lines += [SEP,
              f"  PIPELINE-2 PROBING REPORT — {model_display}",
              f"  Dataset : {dataset_display}",
              f"  Generated : {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}",
              SEP, ""]

    lines.append("  PROBE ANALYSIS  (test = contrastive test only)")
    lines.append(SEP2)
    lines.append(
        f"  {'Layer':>10}  {'AUROC_all':>10}  {'AUROC_cont':>10}  "
        f"{'Gap':>8}  {'Dir':>12}  {'n_R':>6}  {'n_W':>6}"
    )
    lines.append("  " + "─" * 68)

    selected = {r["layer_name"] for r in top_layers}
    for r in all_results:
        flag  = " *" if r["layer_name"] in selected else "  "
        a_all = f"{r['auroc_all']:.3f}"  if r["auroc_all"]  is not None else "   N/A"
        a_cnt = f"{r['auroc_cont']:.3f}" if r["auroc_cont"] is not None else "   N/A"
        gap_s = f"{r['gap']:+.3f}"       if r["gap"]        is not None else "   N/A"
        lines.append(
            f"{flag} {r['layer_name']:>10}  {a_all:>10}  {a_cnt:>10}  "
            f"{gap_s:>8}  {r['dir_source']:>12}  "
            f"{r['n_right_all']:>6}  {r['n_wrong_all']:>6}"
        )
    lines += ["", "  (* = selected for steering)", ""]

    lines.append("  SELECTED TOP LAYERS  (sorted by AUROC_cont descending)")
    lines.append(SEP2)
    for rank, r in enumerate(
        sorted(top_layers, key=lambda x: x.get("auroc_cont") or 0.0, reverse=True), 1
    ):
        ac_s = f"{r['auroc_cont']:.3f}" if r.get("auroc_cont") is not None else "N/A"
        lines.append(
            f"  #{rank}  {r['layer_name']}  AUROC_cont={ac_s}  dir_source={r['dir_source']}"
        )
    lines += ["", SEP]

    text = "\n".join(lines)
    print(text)
    path = probe_dir / "probing_report.txt"
    path.write_text(text, encoding="utf-8")
    print(f"  Report  → {path}")



def write_probing_report_xlsx(
    probe_dir:       Path,
    model_display:   str,
    dataset_display: str,
    all_results:     list[dict],
    top_layers:      list[dict],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WARN] openpyxl not installed — Excel report skipped.")
        return

    wb     = openpyxl.Workbook()
    HEAD_F = PatternFill("solid", fgColor="FFD9E1F2")
    BEST_F = PatternFill("solid", fgColor="FFFFD700")
    SEL_F  = PatternFill("solid", fgColor="FFC6EFCE")
    BOLD   = Font(bold=True)
    CTR    = Alignment(horizontal="center")

    selected = {r["layer_name"] for r in top_layers}

    ws = wb.active
    ws.title = "Probe Analysis"
    ws.append([f"PIPELINE-2 PROBE — {model_display} — {dataset_display}"])
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.append([])

    hdr = ["Layer", "AUROC_all", "AUROC_cont", "Gap", "Dir Source",
           "n_Right", "n_Wrong", "Selected?", "bootstrap_k"]
    ws.append(hdr)
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(3, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
    ws.freeze_panes = "A4"

    best_ac_val = max(
        (r["auroc_cont"] for r in all_results if r["auroc_cont"] is not None),
        default=None,
    )
    for r in all_results:
        is_sel = r["layer_name"] in selected
        row = [
            r["layer_name"], r["auroc_all"], r["auroc_cont"], r["gap"],
            r["dir_source"], r["n_right_all"], r["n_wrong_all"],
            "✓" if is_sel else "", r["bootstrap_k"],
        ]
        ws.append(row)
        ri = ws.max_row
        fill = SEL_F if is_sel else (BEST_F if r.get("auroc_cont") == best_ac_val else None)
        for ci in range(1, len(row) + 1):
            ws.cell(ri, ci).alignment = CTR
            if fill:
                ws.cell(ri, ci).fill = fill

    for col, w in zip("ABCDEFGHI", [12, 12, 12, 8, 14, 9, 9, 10, 12]):
        ws.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Top Layers")
    ws3.append([f"SELECTED TOP-N LAYERS — {model_display}"])
    ws3.cell(1, 1).font = Font(bold=True, size=13)
    ws3.append([])

    hdr3 = ["Layer", "AUROC_cont", "Gap", "Dir Source", "bootstrap_k", "raw_norm"]
    ws3.append(hdr3)
    for ci, h in enumerate(hdr3, 1):
        c = ws3.cell(3, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR

    for r in top_layers:
        ws3.append([
            r["layer_name"], r.get("auroc_cont"), r.get("gap"),
            r.get("dir_source"), r.get("bootstrap_k"), round(r.get("raw_norm", 0), 4),
        ])
        for ci in range(1, 7):
            ws3.cell(ws3.max_row, ci).alignment = CTR

    for col in "ABCDEF":
        ws3.column_dimensions[col].width = 14

    path = probe_dir / "probing_report.xlsx"
    wb.save(path)
    print(f"  Excel   → {path}")



def save_probe_plots(
    probe_dir:     Path,
    model_display: str,
    all_results:   list[dict],
    top_layers:    list[dict],
) -> None:
    plot_dir = probe_dir / "plots"
    plot_dir.mkdir(exist_ok=True)

    layer_idx = [r["layer_idx"]  for r in all_results]
    au_all    = [r["auroc_all"]  for r in all_results]
    au_cont   = [r["auroc_cont"] for r in all_results]
    gaps      = [r["gap"]        for r in all_results]
    top_idx   = {r["layer_idx"]  for r in top_layers}

    def _nz(lst):
        return [v if v is not None else 0.0 for v in lst]

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(layer_idx, _nz(au_all),  "o-", color="steelblue",  lw=1.5, label="AUROC (all-data)")
    ax.plot(layer_idx, _nz(au_cont), "s-", color="darkorange", lw=1.5, label="AUROC (cont-only)")
    for li in top_idx:
        ax.axvline(li, color="gray", linestyle=":", lw=0.9)
    ax.axhline(0.5, color="silver", linestyle="--", lw=0.8)
    ax.set_xlabel("Layer index"); ax.set_ylabel("AUROC")
    ax.set_title(f"Probe AUROC — {model_display}")
    ax.legend(fontsize=9); ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(plot_dir / "probe_auroc.png", dpi=150)
    plt.close(fig)

    gaps_p = _nz(gaps)
    colors = ["tomato" if g > GAP_THRESH else "steelblue" for g in gaps_p]
    fig, ax = plt.subplots(figsize=(10, 3))
    ax.bar(layer_idx, gaps_p, color=colors, width=0.7)
    ax.axhline(GAP_THRESH, color="red", linestyle="--", lw=1, label=f"threshold ({GAP_THRESH})")
    ax.axhline(0, color="black", lw=0.5)
    ax.set_xlabel("Layer index"); ax.set_ylabel("Gap (AUROC_all − AUROC_cont)")
    ax.set_title(f"Probe Gap — {model_display}")
    ax.legend(fontsize=9); ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(plot_dir / "probe_gap.png", dpi=150)
    plt.close(fig)

    print(f"  Plots   → {plot_dir}")



def generate_dataset_probing_report(
    ds_results_dir: Path,
    dataset_key:    str,
    all_model_data: list[dict],
) -> None:
    """
    Aggregate probing results for all models into a dataset-level report.

    Each entry of all_model_data:
      {"model_key", "model_display", "probe_analysis", "top_layers"}
    """
    if not all_model_data:
        return

    ds_display = _DS_DISPLAY.get(dataset_key, dataset_key)
    SEP  = "=" * 80
    SEP2 = "-" * 80
    lines: list[str] = []

    lines += [
        SEP,
        f"  PIPELINE-2 PROBING SUMMARY — {ds_display}",
        f"  Models   : {len(all_model_data)}",
        f"  Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}",
        SEP, "",
    ]

    lines.append("  PROBE AUROC SUMMARY")
    lines.append(SEP2)
    lines.append(
        f"  {'Model':25}  {'Best AUROC_all':>15}  {'Best AUROC_cont':>15}  {'Top Layer':>10}"
    )
    lines.append("  " + "─" * 70)
    for m in all_model_data:
        pa   = m.get("probe_analysis", [])
        tops = m.get("top_layers", [])
        baa  = max((r["auroc_all"]  for r in pa if r.get("auroc_all")  is not None), default=None)
        bac  = max((r["auroc_cont"] for r in pa if r.get("auroc_cont") is not None), default=None)
        tl   = tops[0]["layer_name"] if tops else "N/A"
        lines.append(
            f"  {m['model_display']:25}"
            f"  {f'{baa:.3f}' if baa is not None else '        N/A':>15}"
            f"  {f'{bac:.3f}' if bac is not None else '        N/A':>15}"
            f"  {tl:>10}"
        )
    lines += ["", SEP]

    text = "\n".join(lines)
    print(text)
    txt_path = ds_results_dir / f"{dataset_key}_probing_report.txt"
    txt_path.write_text(text, encoding="utf-8")
    print(f"  Dataset probing report → {txt_path}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return

    wb     = openpyxl.Workbook()
    HEAD_F = PatternFill("solid", fgColor="FFD9E1F2")
    BEST_F = PatternFill("solid", fgColor="FFFFD700")
    BOLD   = Font(bold=True)
    CTR    = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "AUROC Summary"
    ws.append([f"PIPELINE-2 PROBE — {ds_display}"])
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.append([])
    hdr = ["Model", "Best AUROC_all", "Best AUROC_cont", "Best Gap", "Top Layer (first)", "# Top Layers"]
    ws.append(hdr)
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(3, ci, h); c.font = BOLD; c.fill = HEAD_F; c.alignment = CTR
    ws.freeze_panes = "A4"
    for m in all_model_data:
        pa   = m.get("probe_analysis", [])
        tops = m.get("top_layers", [])
        baa  = max((r["auroc_all"]  for r in pa if r.get("auroc_all")  is not None), default=None)
        bac  = max((r["auroc_cont"] for r in pa if r.get("auroc_cont") is not None), default=None)
        gap  = max((r["gap"]        for r in pa if r.get("gap")        is not None), default=None)
        tl   = tops[0]["layer_name"] if tops else "N/A"
        ws.append([m["model_display"], baa, bac, gap, tl, len(tops)])
        ri = ws.max_row
        if bac is not None and bac == max(
            (r2.get("auroc_cont", 0) or 0 for m2 in all_model_data for r2 in m2.get("probe_analysis", [])),
            default=0,
        ):
            ws.cell(ri, 3).fill = BEST_F
        for ci in range(1, 7):
            ws.cell(ri, ci).alignment = CTR
    ws.column_dimensions["A"].width = 28
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16

    xlsx_path = ds_results_dir / f"{dataset_key}_probing_report.xlsx"
    wb.save(xlsx_path)
    print(f"  Dataset Excel report  → {xlsx_path}")



def run_probing(
    model_key:   str,
    acts_dir:    str,
    results_dir: Path,
    dataset_key: str,
    top_n:       int   = TOP_N_LAYERS,
    gap_thresh:  float = GAP_THRESH,
    n_bootstrap: int   = N_BOOTSTRAP,
    seed:        int   = RNG_SEED,
) -> dict | None:
    """
    Run full probe for one model.

    Parameters
    ----------
    acts_dir    : activation run directory (contains H5 files + split.json)
    results_dir : pipeline2 output directory for this model run
    dataset_key : dataset name (for labeling)

    Returns
    -------
    dict with "probe_analysis", "top_layers" (for dataset summary)
    """
    model_cfg  = MODEL_REGISTRY[model_key]
    probe_dir  = results_dir / "probing"
    done_flag  = probe_dir / "probe_analysis.json"

    if done_flag.exists():
        print(f"  [SKIP] {model_cfg['display']} — probing already done.")
        with open(done_flag) as f:
            probe_data = json.load(f)
        tl_path  = probe_dir / "top_layers.json"
        top_lays = json.load(open(tl_path)) if tl_path.exists() else []
        return {
            "model_key":      model_key,
            "model_display":  model_cfg["display"],
            "probe_analysis": probe_data.get("all_layers", []),
            "top_layers":     top_lays,
        }

    probe_dir.mkdir(parents=True, exist_ok=True)
    ds_display = _DS_DISPLAY.get(dataset_key, dataset_key)

    print(f"\n{'='*72}")
    print(f"  PIPELINE-2 PROBING — {model_cfg['display']}  [{ds_display}]")
    print(f"  top_n={top_n}  gap_thresh={gap_thresh}"
          f"  n_bootstrap={n_bootstrap}  seed={seed}")
    print(f"  acts_dir : {acts_dir}")
    print(f"  out_dir  : {results_dir}")
    print("="*72)

    split_path = Path(acts_dir) / "split.json"
    if not split_path.exists():
        print(f"  [ERROR] split.json not found in {acts_dir}")
        return None

    with open(split_path) as f:
        split = json.load(f)

    try:
        all_results, top_layers = analyze_layers(
            acts_dir, split,
            top_n=top_n, gap_thresh=gap_thresh,
            n_bootstrap=n_bootstrap, seed=seed,
        )
    except RuntimeError as exc:
        print(f"  [ERROR] analyze_layers failed: {exc}")
        return None

    dir_dir = probe_dir / "directions"
    dir_dir.mkdir(exist_ok=True)
    for rec in top_layers:
        np.save(dir_dir / f"{rec['layer_name']}.npy", rec["_direction"])

    def _clean(r: dict) -> dict:
        return {k: v for k, v in r.items() if k != "_direction"}

    probe_data = {
        "model":       model_key,
        "dataset":     dataset_key,
        "acts_dir":    str(acts_dir),
        "top_n":       top_n,
        "gap_thresh":  gap_thresh,
        "n_bootstrap": n_bootstrap,
        "seed":        seed,
        "all_layers":  [_clean(r) for r in all_results],
    }
    top_layers_json = [_clean(r) for r in top_layers]

    with open(done_flag, "w") as f:
        json.dump(probe_data, f, indent=2)
    with open(probe_dir / "top_layers.json", "w") as f:
        json.dump(top_layers_json, f, indent=2)

    print(f"\n  {'Layer':>10}  {'AUROC_all':>10}  {'AUROC_cont':>10}"
          f"  {'Gap':>8}  {'Dir':>12}  {'n_R':>6}  {'n_W':>6}")
    print("  " + "-"*68)
    sel_names = {r["layer_name"] for r in top_layers}
    for r in all_results:
        flag  = " *" if r["layer_name"] in sel_names else "  "
        a_all = f"{r['auroc_all']:.3f}"  if r["auroc_all"]  is not None else "  N/A "
        a_cnt = f"{r['auroc_cont']:.3f}" if r["auroc_cont"] is not None else "  N/A "
        gap_s = f"{r['gap']:+.3f}"       if r["gap"]        is not None else "  N/A "
        print(f"{flag} {r['layer_name']:>10}  {a_all:>10}  {a_cnt:>10}"
              f"  {gap_s:>8}  {r['dir_source']:>12}"
              f"  {r['n_right_all']:>6}  {r['n_wrong_all']:>6}")
    print(f"\n  (* = selected for steering)")

    write_probing_report_txt(
        probe_dir, model_cfg["display"], ds_display,
        all_results, top_layers,
    )
    write_probing_report_xlsx(
        probe_dir, model_cfg["display"], ds_display,
        all_results, top_layers,
    )
    try:
        save_probe_plots(probe_dir, model_cfg["display"],
                         all_results, top_layers_json)
    except Exception as exc:
        print(f"  [WARN] Plotting failed: {exc}")

    print(f"\n  Saved → {probe_dir}")

    return {
        "model_key":      model_key,
        "model_display":  model_cfg["display"],
        "probe_analysis": [_clean(r) for r in all_results],
        "top_layers":     top_layers_json,
    }



def main() -> None:
    parser = argparse.ArgumentParser(
        description="Pipeline-2 probing: probe analysis on stored activations."
    )
    parser.add_argument("--dataset",      default="mbppplus", choices=["mbppplus"])
    parser.add_argument("--model",        default=None, choices=list(MODEL_REGISTRY.keys()))
    parser.add_argument("--acts-dir",     default=None,
                        help="Activation run directory (has split.json + H5 files).")
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT,
                        help="Root for probing results (default: ../mbppplus).")
    parser.add_argument("--base",         default=DEFAULT_BASE,
                        help="Root of activation dirs (default: ../mbppplus).")
    parser.add_argument("--top-n",        type=int,   default=TOP_N_LAYERS)
    parser.add_argument("--n-bootstrap",  type=int,   default=N_BOOTSTRAP)
    parser.add_argument("--gap-thresh",   type=float, default=GAP_THRESH)
    parser.add_argument("--rng-seed",     type=int,   default=RNG_SEED)
    args = parser.parse_args()

    results_root = Path(args.results_root).expanduser()
    kw = dict(
        top_n=args.top_n, gap_thresh=args.gap_thresh,
        n_bootstrap=args.n_bootstrap, seed=args.rng_seed,
    )

    if args.acts_dir and args.model and args.dataset:
        acts_dir    = args.acts_dir
        existing    = _find_p2_run(results_root, args.dataset, args.model,
                                   require_file="probing/probe_analysis.json")
        if existing:
            results_dir = existing
            print(f"  [RESUME] Using existing results dir: {results_dir}")
        else:
            results_dir = _new_results_dir(results_root, args.dataset, args.model)
        run_probing(
            model_key=args.model, acts_dir=acts_dir,
            results_dir=results_dir, dataset_key=args.dataset, **kw,
        )
        return

    base = Path(args.base).expanduser()
    if not base.exists():
        print(f"[ERROR] Base activations dir not found: {base}")
        return

    ds_key = "mbppplus"
    runs   = _find_activation_runs(base)

    print("\n" + "=" * 66)
    print("  CODE_LLM PROBING — MBPP+ ACTIVATION RUNS")
    print("=" * 66)
    if runs:
        for i, (mk, rd) in enumerate(runs, 1):
            label = MODEL_REGISTRY[mk]["display"] if mk in MODEL_REGISTRY else mk
            print(f"  {i}. {label:<28}  ←  {rd.name}")
    else:
        print("  No activation runs found.")
        print("  Run: python contrastive_set_gen.py")
    print("=" * 66)

    if not runs:
        return

    ds_dir   = base
    available = [(ds_key, ds_dir, runs)]

    ds_results_dir = results_root / ds_key
    ds_results_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n  {len(runs)} model(s) in {_DS_DISPLAY.get(ds_key, ds_key)}:")
    for i, (mk, rd) in enumerate(runs, 1):
        label = MODEL_REGISTRY[mk]["display"] if mk in MODEL_REGISTRY else mk
        print(f"    {i}. {label}  \u2190  {rd.name}")

    print(f"\n  Select model(s) to probe [1-{len(runs)}, comma-separated, or Enter for all]: ", end="", flush=True)
    raw = input().strip()
    if raw:
        selected_indices = []
        for part in raw.split(","):
            part = part.strip()
            if part.isdigit() and 1 <= int(part) <= len(runs):
                selected_indices.append(int(part) - 1)
            else:
                print(f"  [WARN] Ignoring invalid selection: {part!r}")
        if selected_indices:
            runs = [runs[i] for i in selected_indices]
        else:
            print("  No valid selection — running all models.")
    print(f"  Running probing for: {', '.join(MODEL_REGISTRY[mk]['display'] if mk in MODEL_REGISTRY else mk for mk, _ in runs)}")

    all_model_data: list[dict] = []

    for i, (mk, acts_dir) in enumerate(runs, 1):
        print(f"\n{'#'*72}")
        print(f"  Model {i}/{len(runs)}: {MODEL_REGISTRY.get(mk, {}).get('display', mk)}")
        print(f"{'#'*72}")

        existing = _find_p2_run(results_root, ds_key, mk,
                                require_file="probing/probe_analysis.json")
        if existing:
            results_dir = existing
            print(f"  [RESUME] Found existing results: {results_dir}")
        else:
            results_dir = _new_results_dir(results_root, ds_key, mk)

        try:
            data = run_probing(
                model_key=mk, acts_dir=str(acts_dir),
                results_dir=results_dir, dataset_key=ds_key, **kw,
            )
            if data:
                all_model_data.append(data)
        except Exception as exc:
            import traceback
            print(f"\n  [ERROR] {MODEL_REGISTRY.get(mk, {}).get('display', mk)} failed: {exc}")
            traceback.print_exc()

    if all_model_data:
        print(f"\n{'='*72}")
        print("  Generating dataset probing report \u2026")
        generate_dataset_probing_report(ds_results_dir, ds_key, all_model_data)


if __name__ == "__main__":
    main()
